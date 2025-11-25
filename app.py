"""
Aplicativo Flask simples (PoC) para gerenciar Projetos, Macroetapas, Etapas e Tarefas,
calculando datas automaticamente de baixo para cima.

Execução:
    - Instale dependências: pip install flask flask_sqlalchemy
    - Rode: python app.py
    - Acesse: http://127.0.0.1:5000
"""

from datetime import datetime, date, timedelta
from urllib.parse import quote
from flask import Flask, render_template, redirect, url_for, request, abort, send_file, make_response, flash
from sqlalchemy import or_, and_
from sqlalchemy.orm import joinedload
from models import db, Project, MacroStage, Stage, Task, WeeklyUpdate
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import re
import os

app = Flask(__name__)
# Configuração do banco SQLite usando caminho absoluto baseado no diretório do projeto
basedir = os.path.abspath(os.path.dirname(__file__))
db_path = os.path.join(basedir, "instance", "schedule.db")
app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{db_path}"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SECRET_KEY"] = "dev-secret-key-change-in-production"  # Necessário para flash messages

# Inicializa o objeto db com o app
db.init_app(app)


PROJECT_STATUS_CHOICES = [
    "A iniciar",
    "Em andamento",
    "Concluído",
    "Suspenso",
    "Descartado",
]

MACRO_STRUCTURE_CHOICES = {"stages", "tasks"}
STAGE_TYPE_CHOICES = ["robô", "sistema", "não se aplica"]


def redirect_with_anchor(endpoint: str, anchor: str, **values):
    url = url_for(endpoint, **values)
    if anchor:
        url = f"{url}#{anchor}"
    return redirect(url)


def task_parent_context(task: Task):
    if task.stage:
        return task.stage.macrostage.project.id, f"stage-{task.stage.id}"
    return task.macrostage.project.id, f"macrostage-{task.macrostage.id}"


def parse_date_field(raw_value: str):
    """
    Converte uma string no formato YYYY-MM-DD em objeto date.
    Retorna None caso não informado ou inválido.
    """
    if not raw_value:
        return None

    try:
        return datetime.strptime(raw_value, "%Y-%m-%d").date()
    except ValueError:
        return None


def validate_task_dates(start_date, end_date):
    """
    Valida que a data de início nunca seja maior que a data de fim.
    As datas podem ser iguais (coincidir).
    
    Args:
        start_date: Data de início (date ou None)
        end_date: Data de fim (date ou None)
    
    Returns:
        tuple: (is_valid: bool, error_message: str ou None)
    """
    # Se ambas as datas estão presentes, valida
    if start_date is not None and end_date is not None:
        if start_date > end_date:
            return False, "A data de início não pode ser maior que a data de fim."
        if end_date < start_date:
            return False, "A data de fim não pode ser menor que a data de início."
    
    return True, None


def calculate_project_progress(start_date, end_date, current_date=None):
    """
    Calcula o progresso temporal de um projeto em porcentagem.
    
    Args:
        start_date: Data de início do projeto (date ou None)
        end_date: Data de fim do projeto (date ou None)
        current_date: Data atual para cálculo (date, padrão: date.today())
    
    Returns:
        int: Porcentagem de progresso (0-100) ou None se não puder calcular
    """
    if start_date is None or end_date is None:
        return None
    
    if current_date is None:
        current_date = date.today()
    
    # Se a data atual é anterior ao início, retorna 0%
    if current_date < start_date:
        return 0
    
    # Se a data atual é posterior ao fim, retorna 100%
    if current_date > end_date:
        return 100
    
    # Se início e fim são iguais, retorna 100%
    if start_date == end_date:
        return 100
    
    # Calcula o progresso: (atual - início) / (fim - início) * 100
    total_days = (end_date - start_date).days
    elapsed_days = (current_date - start_date).days
    
    if total_days == 0:
        return 100
    
    progress = int((elapsed_days / total_days) * 100)
    
    # Garante que está entre 0 e 100
    return max(0, min(100, progress))


def calculate_automatic_status(project: Project) -> str:
    """
    Calcula o status automático de um projeto baseado em suas datas de início e fim.
    
    Regras:
        - "A iniciar": start_date é None (sem tarefas) OU start_date > hoje
        - "Em andamento": start_date <= hoje <= end_date (projeto iniciou e ainda não terminou)
        - "Concluído": end_date não é None E end_date < hoje (projeto já terminou)
    
    Args:
        project: Projeto para calcular o status
    
    Returns:
        str: Status calculado ("A iniciar", "Em andamento" ou "Concluído")
    """
    today = date.today()
    
    # Se não há data de início (sem tarefas), está "a iniciar"
    if project.start_date is None:
        return "A iniciar"
    
    # Se a data de início é no futuro, está "a iniciar"
    if project.start_date > today:
        return "A iniciar"
    
    # Se há data de fim e já passou, está "concluído"
    if project.end_date is not None and project.end_date < today:
        return "Concluído"
    
    # Se está entre início e fim (ou sem fim definido mas já iniciou), está "em andamento"
    if project.start_date <= today:
        if project.end_date is None or project.end_date >= today:
            return "Em andamento"
    
    # Fallback: se chegou aqui, considera "em andamento"
    return "Em andamento"


def get_project_status(project: Project) -> dict:
    """
    Retorna o status efetivo de um projeto (manual ou automático).
    
    Args:
        project: Projeto para obter o status
    
    Returns:
        dict: {
            'value': str - status atual (manual ou automático),
            'is_manual': bool - indica se é manual,
            'display_text': str - texto formatado para exibição
        }
    """
    if project.status_manual and project.status_manual_value:
        # Status manual
        return {
            'value': project.status_manual_value,
            'is_manual': True,
            'display_text': f"{project.status_manual_value}"
        }
    else:
        # Status automático
        automatic_status = calculate_automatic_status(project)
        return {
            'value': automatic_status,
            'is_manual': False,
            'display_text': automatic_status
        }


def calculate_stage_status(stage: Stage) -> str:
    """
    Calcula o status de uma etapa (robô/sistema) baseado nas tarefas associadas.
    
    Regras:
        - "A iniciar": nenhuma tarefa OU todas as tarefas têm start_date > hoje
        - "Em andamento": pelo menos uma tarefa com start_date ≤ hoje E 
          ainda existe pelo menos uma tarefa futura (start_date > hoje OU end_date > hoje)
        - "Concluído": todas as tarefas têm end_date ≤ hoje
    
    Args:
        stage: Etapa para calcular o status
    
    Returns:
        str: Status calculado ("A iniciar", "Em andamento" ou "Concluído")
    """
    today = date.today()
    tasks = stage.tasks
    
    # Caso B: Nenhuma tarefa ou todas as tarefas começam no futuro
    if not tasks:
        return "A iniciar"
    
    # Coleta todas as datas de início e fim válidas
    start_dates = [t.start_date for t in tasks if t.start_date is not None]
    end_dates = [t.end_date for t in tasks if t.end_date is not None]
    
    # Se não há datas de início válidas, considera "A iniciar"
    if not start_dates:
        return "A iniciar"
    
    # Se todas as tarefas começam no futuro
    if all(sd > today for sd in start_dates):
        return "A iniciar"
    
    # Caso D: Todas as tarefas terminaram (todas têm end_date e todas são <= hoje)
    if end_dates and len(end_dates) == len(tasks) and all(ed <= today for ed in end_dates):
        return "Concluído"
    
    # Caso C: Pelo menos uma tarefa já começou E ainda há tarefas futuras
    has_started = any(sd <= today for sd in start_dates)
    
    # Verifica se ainda há tarefas futuras
    # Uma tarefa é futura se: start_date > hoje OU end_date > hoje OU (start_date <= hoje mas end_date é None)
    has_future = False
    for task in tasks:
        # Tarefa futura se começa no futuro
        if task.start_date and task.start_date > today:
            has_future = True
            break
        # Tarefa futura se termina no futuro
        if task.end_date and task.end_date > today:
            has_future = True
            break
        # Tarefa futura se já começou mas não tem data de fim (ainda em andamento)
        if task.start_date and task.start_date <= today and task.end_date is None:
            has_future = True
            break
    
    # Se pelo menos uma começou e ainda há tarefas futuras, está "Em andamento"
    if has_started and has_future:
        return "Em andamento"
    
    # Se pelo menos uma começou mas não há tarefas futuras, todas terminaram
    if has_started:
        # Se todas as tarefas têm end_date e todas terminaram
        if end_dates and len(end_dates) == len(tasks) and all(ed <= today for ed in end_dates):
            return "Concluído"
        # Caso contrário, ainda está em andamento
        return "Em andamento"
    
    # Fallback: considera "Em andamento"
    return "Em andamento"


# --------------------------
# Funções auxiliares de datas
# --------------------------

def recalculate_stage(stage: Stage) -> None:
    """
    Recalcula as datas de início e fim de uma Etapa (Stage) com base nas suas Tarefas.

    Regra:
        - start_date = menor data de início entre as tarefas
        - end_date   = maior data de fim entre as tarefas
        - se não houver tarefas com datas válidas, ambos ficam como None
    """
    if stage is None:
        return

    tasks = stage.tasks

    start_dates = [t.start_date for t in tasks if t.start_date is not None]
    end_dates = [t.end_date for t in tasks if t.end_date is not None]

    stage.start_date = min(start_dates) if start_dates else None
    stage.end_date = max(end_dates) if end_dates else None


def recalculate_macrostage(macrostage: MacroStage) -> None:
    """
    Recalcula as datas de início e fim de uma Macroetapa com base nas Etapas.

    Regra:
        - start_date = menor start_date entre as etapas
        - end_date   = maior end_date entre as etapas
        - se não houver etapas com datas válidas, ambos ficam como None
    """
    if macrostage is None:
        return

    stages = macrostage.stages
    tasks = macrostage.tasks

    start_dates = [s.start_date for s in stages if s.start_date is not None]
    end_dates = [s.end_date for s in stages if s.end_date is not None]

    task_starts = [t.start_date for t in tasks if t.start_date is not None]
    task_ends = [t.end_date for t in tasks if t.end_date is not None]

    start_dates.extend(task_starts)
    end_dates.extend(task_ends)

    macrostage.start_date = min(start_dates) if start_dates else None
    macrostage.end_date = max(end_dates) if end_dates else None


def recalculate_project(project: Project) -> None:
    """
    Recalcula as datas de início e fim de um Projeto com base nas Macroetapas.

    Regra:
        - start_date = menor start_date entre as macroetapas
        - end_date   = maior end_date entre as macroetapas
        - se não houver macroetapas com datas válidas, ambos ficam como None
    """
    if project is None:
        return

    macrostages = project.macrostages

    start_dates = [m.start_date for m in macrostages if m.start_date is not None]
    end_dates = [m.end_date for m in macrostages if m.end_date is not None]

    project.start_date = min(start_dates) if start_dates else None
    project.end_date = max(end_dates) if end_dates else None


def recalculate_project_status(project: Project) -> None:
    """
    Recalcula o status automático de um projeto.
    Só atualiza se o status não for manual.
    """
    if project is None:
        return
    
    # Só recalcula se não for status manual
    if not project.status_manual:
        automatic_status = calculate_automatic_status(project)
        project.status = automatic_status


def recalculate_all_from_stage(stage: Stage) -> None:
    """
    A partir de uma Etapa (Stage), recalcula em cadeia:
        - Stage (etapa)
        - MacroStage (macroetapa)
        - Project (projeto)
    e realiza o commit das alterações.
    """
    if stage is None:
        return

    macrostage = stage.macrostage
    project = macrostage.project if macrostage else None

    # Recalcula em cascata
    recalculate_stage(stage)
    recalculate_macrostage(macrostage)
    recalculate_project(project)
    recalculate_project_status(project)

    db.session.commit()


def recalculate_all_from_macrostage(macrostage: MacroStage) -> None:
    if macrostage is None:
        return

    project = macrostage.project

    recalculate_macrostage(macrostage)
    recalculate_project(project)
    recalculate_project_status(project)

    db.session.commit()


# --------------------------
# Funções auxiliares para ajuste automático de tarefas
# --------------------------

def calculate_task_shift_delta(old_start_date, old_end_date, new_start_date, new_end_date):
    """
    Calcula o delta de dias de deslocamento de uma tarefa.
    
    Regras:
        - Prioridade: end_date > start_date
        - Se end_date mudou: delta_days = (new_end_date - old_end_date).days
        - Senão, se start_date mudou: delta_days = (new_start_date - old_start_date).days
        - Se ambos mudaram mas delta é diferente, usar end_date
        - Retorna (delta_days, old_start_date) - sempre usa start_date como referência para tarefas subsequentes
    
    Args:
        old_start_date: Data de início antiga (date ou None)
        old_end_date: Data de fim antiga (date ou None)
        new_start_date: Data de início nova (date ou None)
        new_end_date: Data de fim nova (date ou None)
    
    Returns:
        tuple: (delta_days: int, reference_start_date: date ou None)
    """
    # Se end_date mudou, calcula delta baseado em end_date
    if old_end_date is not None and new_end_date is not None:
        delta_end = (new_end_date - old_end_date).days
        if delta_end != 0:
            # Usa old_start_date como referência (tarefas subsequentes são baseadas em start_date)
            return delta_end, old_start_date
    
    # Se end_date não mudou ou não existe, verifica start_date
    if old_start_date is not None and new_start_date is not None:
        delta_start = (new_start_date - old_start_date).days
        if delta_start != 0:
            # Usa start_date antiga como referência
            return delta_start, old_start_date
    
    # Se ambos mudaram mas delta é diferente, prioriza end_date para o cálculo do delta
    if (old_end_date is not None and new_end_date is not None and
        old_start_date is not None and new_start_date is not None):
        delta_end = (new_end_date - old_end_date).days
        delta_start = (new_start_date - old_start_date).days
        if delta_end != 0 and delta_start != 0 and delta_end != delta_start:
            # Usa delta_end mas old_start_date como referência
            return delta_end, old_start_date
    
    # Nenhum deslocamento detectado
    return 0, None


def find_subsequent_tasks(project_id, reference_start_date, exclude_task_id):
    """
    Encontra todas as tarefas subsequentes de um projeto que devem ser deslocadas.
    
    Uma tarefa é considerada subsequente se:
        - Pertence ao mesmo projeto
        - start_date > reference_start_date (data de início da tarefa editada)
        - Não é a tarefa excluída (exclude_task_id)
    
    Args:
        project_id: ID do projeto
        reference_start_date: Data de início de referência (date) - data de início antiga da tarefa editada
        exclude_task_id: ID da tarefa a excluir da busca
    
    Returns:
        list: Lista de tarefas ordenadas por start_date e position
    """
    if reference_start_date is None:
        return []
    
    # Busca todas as tarefas do projeto com start_date > reference_start_date
    tasks = Task.query.join(MacroStage).filter(
        MacroStage.project_id == project_id,
        Task.start_date.isnot(None),
        Task.start_date > reference_start_date,
        Task.id != exclude_task_id
    ).order_by(Task.start_date, Task.position).all()
    
    return tasks


# --------------------------
# Funções auxiliares para Excel
# --------------------------

def format_excel_header(worksheet, row, num_cols):
    """
    Formata a linha de cabeçalho do Excel.
    
    Args:
        worksheet: Planilha do openpyxl
        row: Número da linha (1-indexed)
        num_cols: Número de colunas a formatar
    """
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in range(1, num_cols + 1):
        cell = worksheet.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border


def format_excel_date(cell, date_value):
    """
    Formata uma célula com data no formato brasileiro (dd/mm/yyyy).
    
    Args:
        cell: Célula do openpyxl
        date_value: Objeto date ou None
    """
    if date_value:
        cell.value = date_value.strftime("%d/%m/%Y")
        cell.number_format = "DD/MM/YYYY"
    else:
        cell.value = "—"


def auto_adjust_column_width(worksheet):
    """
    Ajusta automaticamente a largura das colunas baseado no conteúdo.
    
    Args:
        worksheet: Planilha do openpyxl
    """
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        # Ajusta largura com margem
        adjusted_width = min(max_length + 2, 50)  # Máximo de 50 caracteres
        worksheet.column_dimensions[column_letter].width = adjusted_width


def create_excel_response(workbook, filename):
    """
    Cria uma resposta HTTP para download do arquivo Excel.
    
    Args:
        workbook: Workbook do openpyxl
        filename: Nome do arquivo (sem extensão .xlsx)
    
    Returns:
        Response do Flask para download
    """
    # Gera arquivo em memória
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    
    # Cria resposta HTTP
    response = make_response(send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"{filename}.xlsx"
    ))
    
    return response


def sanitize_filename(name):
    """
    Remove caracteres inválidos de um nome de arquivo.
    
    Args:
        name: Nome original
    
    Returns:
        Nome sanitizado
    """
    # Remove caracteres inválidos para nomes de arquivo
    sanitized = re.sub(r'[<>:"/\\|?*]', '_', name)
    # Remove espaços múltiplos
    sanitized = re.sub(r'\s+', '_', sanitized)
    # Remove underscores múltiplos
    sanitized = re.sub(r'_+', '_', sanitized)
    # Remove underscores no início e fim
    sanitized = sanitized.strip('_')
    return sanitized


# --------------
# Rotas principais
# --------------

@app.route("/")
def index():
    """
    Redireciona para a listagem de projetos.
    """
    return redirect(url_for("list_projects"))


@app.route("/projects", methods=["GET"])
def list_projects():
    """
    Lista todos os projetos cadastrados, com suas datas calculadas.
    """
    projects = Project.query.order_by(Project.id).all()
    return render_template("projects.html", projects=projects)


@app.route("/dashboard/projects", methods=["GET"])
def dashboard_projects():
    """
    Dashboard geral de projetos com filtros avançados.
    Permite filtrar por status, órgão demandante, setor interno, coordenador,
    equipe Automatiza, gestor responsável, tipo de projeto (robô/sistema) e ferramentas.
    """
    # Lê parâmetros de filtro da query string (suporta múltiplos valores)
    status_filters = [s.strip() for s in request.args.getlist("status") if s.strip()]
    requesting_agency_filters = [a.strip() for a in request.args.getlist("requesting_agency") if a.strip()]
    internal_department_filters = [d.strip() for d in request.args.getlist("internal_department") if d.strip()]
    coordinator_filters = [c.strip() for c in request.args.getlist("coordinator") if c.strip()]
    automation_support_filters = [a.strip() for a in request.args.getlist("automation_support") if a.strip()]
    sponsoring_manager_filters = [m.strip() for m in request.args.getlist("sponsoring_manager") if m.strip()]
    project_type_filters = [p.strip() for p in request.args.getlist("project_type") if p.strip()]
    # Compatibilidade: se não houver lista, tenta valor único (ignora "all" que significa "todos")
    if not project_type_filters:
        single_project_type = request.args.get("project_type", "").strip()
        if single_project_type and single_project_type != "all":
            project_type_filters = [single_project_type]
    tools_filter = request.args.get("tools", "").strip() or None
    progress_sort = request.args.get("progress_sort", "").strip() or None

    # Inicia query base
    query = Project.query

    # Aplica filtros diretos em Project (multi-select)
    if requesting_agency_filters:
        query = query.filter(Project.requesting_agency.in_(requesting_agency_filters))
    if internal_department_filters:
        query = query.filter(Project.internal_department.in_(internal_department_filters))
    if coordinator_filters:
        query = query.filter(Project.coordinator.in_(coordinator_filters))
    if automation_support_filters:
        query = query.filter(Project.automation_support.in_(automation_support_filters))
    if sponsoring_manager_filters:
        query = query.filter(Project.sponsoring_manager.in_(sponsoring_manager_filters))

    # Filtro por tipo de projeto (robô/sistema) - multi-select
    if project_type_filters:
        # Subquery para projetos com etapas do tipo "robô"
        robot_projects = db.session.query(Project.id).join(MacroStage).join(Stage).filter(
            Stage.stage_type == "robô"
        ).distinct().subquery()

        # Subquery para projetos com etapas do tipo "sistema"
        system_projects = db.session.query(Project.id).join(MacroStage).join(Stage).filter(
            Stage.stage_type == "sistema"
        ).distinct().subquery()

        # Lógica para múltiplos tipos selecionados
        project_type_conditions = []
        
        if "robot" in project_type_filters:
            project_type_conditions.append(Project.id.in_(db.session.query(robot_projects.c.id)))
        if "system" in project_type_filters:
            project_type_conditions.append(Project.id.in_(db.session.query(system_projects.c.id)))
        if "both" in project_type_filters:
            # Projetos que aparecem em ambas as subqueries
            project_type_conditions.append(
                and_(
                    Project.id.in_(db.session.query(robot_projects.c.id)),
                    Project.id.in_(db.session.query(system_projects.c.id))
                )
            )
        if "none" in project_type_filters:
            # Projetos que não aparecem em nenhuma das subqueries
            project_type_conditions.append(
                and_(
                    ~Project.id.in_(db.session.query(robot_projects.c.id)),
                    ~Project.id.in_(db.session.query(system_projects.c.id))
                )
            )
        
        # Se houver condições, aplica com OR (projeto pode atender qualquer uma)
        if project_type_conditions:
            if len(project_type_conditions) == 1:
                query = query.filter(project_type_conditions[0])
            else:
                query = query.filter(or_(*project_type_conditions))

    # Filtro por ferramentas/sistemas
    if tools_filter:
        # Usa subquery para evitar conflitos com outros filtros
        tools_projects = db.session.query(Project.id).join(MacroStage).join(Stage).filter(
            or_(
                Stage.tools.like(f"%{tools_filter}%"),
                Stage.other_tools.like(f"%{tools_filter}%")
            )
        ).distinct().subquery()
        query = query.filter(Project.id.in_(db.session.query(tools_projects.c.id)))

    # Ordena e executa a query
    projects = query.order_by(Project.id).all()

    # Coleta valores distintos para popular os selects dos filtros
    # Para status, calcula o status efetivo de todos os projetos para ter a lista completa
    # (não usa o campo status do banco, pois pode estar desatualizado)
    all_projects_for_status = Project.query.all()
    calculated_statuses = set()
    for p in all_projects_for_status:
        status_info = get_project_status(p)
        calculated_statuses.add(status_info['value'])
    
    filter_options = {
        "statuses": sorted(list(calculated_statuses)),
        "requesting_agencies": sorted([
            a[0] for a in Project.query.with_entities(Project.requesting_agency).distinct()
            .filter(Project.requesting_agency.isnot(None)).all()
        ]),
        "internal_departments": sorted([
            d[0] for d in Project.query.with_entities(Project.internal_department).distinct()
            .filter(Project.internal_department.isnot(None)).all()
        ]),
        "coordinators": sorted([
            c[0] for c in Project.query.with_entities(Project.coordinator).distinct()
            .filter(Project.coordinator.isnot(None)).all()
        ]),
        "automation_supports": sorted([
            a[0] for a in Project.query.with_entities(Project.automation_support).distinct()
            .filter(Project.automation_support.isnot(None)).all()
        ]),
        "sponsoring_managers": sorted([
            m[0] for m in Project.query.with_entities(Project.sponsoring_manager).distinct()
            .filter(Project.sponsoring_manager.isnot(None)).all()
        ]),
    }

    # Calcula o status efetivo e progresso para cada projeto (garantindo consistência)
    projects_with_status = []
    for project in projects:
        project_status = get_project_status(project)
        # Calcula progresso usando a mesma lógica de project_detail()
        # Se status for manual, não calcular progresso
        if project_status['is_manual']:
            progresso_percentual = None
        else:
            progresso_percentual = calculate_project_progress(project.start_date, project.end_date)
        
        projects_with_status.append({
            'project': project,
            'status': project_status['value'],
            'status_display': project_status['display_text'],
            'progress': progresso_percentual
        })
    
    # Aplica filtro de status após calcular status efetivo (multi-select)
    if status_filters:
        projects_with_status = [
            item for item in projects_with_status 
            if item['status'] in status_filters
        ]
    
    # Passa os valores atuais dos filtros para manter selecionados no template (como listas)
    current_filters = {
        "status": status_filters,
        "requesting_agency": requesting_agency_filters,
        "internal_department": internal_department_filters,
        "coordinator": coordinator_filters,
        "automation_support": automation_support_filters,
        "sponsoring_manager": sponsoring_manager_filters,
        "project_type": project_type_filters,
        "tools": tools_filter or "",
        "progress_sort": progress_sort or "",
    }
    
    # Aplica ordenação por progresso se solicitado
    if progress_sort in ("asc", "desc"):
        # Ordena por progresso, tratando None como -1 (vai para o final em desc, início em asc)
        def get_sort_key(item):
            progress = item['progress']
            if progress is None:
                return -1 if progress_sort == "desc" else 999
            return progress
        
        projects_with_status.sort(key=get_sort_key, reverse=(progress_sort == "desc"))

    return render_template(
        "dashboard_projects.html",
        projects=projects_with_status,
        filter_options=filter_options,
        current_filters=current_filters,
    )


@app.route("/dashboard/projects/export", methods=["GET"])
def export_dashboard_projects():
    """
    Exporta o dashboard de projetos para Excel (.xlsx).
    Reutiliza toda a lógica de filtros de dashboard_projects().
    """
    # Lê parâmetros de filtro da query string (suporta múltiplos valores - mesma lógica do dashboard)
    status_filters = [s.strip() for s in request.args.getlist("status") if s.strip()]
    requesting_agency_filters = [a.strip() for a in request.args.getlist("requesting_agency") if a.strip()]
    internal_department_filters = [d.strip() for d in request.args.getlist("internal_department") if d.strip()]
    coordinator_filters = [c.strip() for c in request.args.getlist("coordinator") if c.strip()]
    automation_support_filters = [a.strip() for a in request.args.getlist("automation_support") if a.strip()]
    sponsoring_manager_filters = [m.strip() for m in request.args.getlist("sponsoring_manager") if m.strip()]
    project_type_filters = [p.strip() for p in request.args.getlist("project_type") if p.strip()]
    # Compatibilidade: se não houver lista, tenta valor único (ignora "all" que significa "todos")
    if not project_type_filters:
        single_project_type = request.args.get("project_type", "").strip()
        if single_project_type and single_project_type != "all":
            project_type_filters = [single_project_type]
    tools_filter = request.args.get("tools", "").strip() or None
    progress_sort = request.args.get("progress_sort", "").strip() or None

    # Reutiliza toda a lógica de filtros do dashboard_projects
    query = Project.query

    if requesting_agency_filters:
        query = query.filter(Project.requesting_agency.in_(requesting_agency_filters))
    if internal_department_filters:
        query = query.filter(Project.internal_department.in_(internal_department_filters))
    if coordinator_filters:
        query = query.filter(Project.coordinator.in_(coordinator_filters))
    if automation_support_filters:
        query = query.filter(Project.automation_support.in_(automation_support_filters))
    if sponsoring_manager_filters:
        query = query.filter(Project.sponsoring_manager.in_(sponsoring_manager_filters))

    # Filtro por tipo de projeto (robô/sistema) - multi-select
    if project_type_filters:
        robot_projects = db.session.query(Project.id).join(MacroStage).join(Stage).filter(
            Stage.stage_type == "robô"
        ).distinct().subquery()
        system_projects = db.session.query(Project.id).join(MacroStage).join(Stage).filter(
            Stage.stage_type == "sistema"
        ).distinct().subquery()

        # Lógica para múltiplos tipos selecionados
        project_type_conditions = []
        
        if "robot" in project_type_filters:
            project_type_conditions.append(Project.id.in_(db.session.query(robot_projects.c.id)))
        if "system" in project_type_filters:
            project_type_conditions.append(Project.id.in_(db.session.query(system_projects.c.id)))
        if "both" in project_type_filters:
            # Projetos que aparecem em ambas as subqueries
            project_type_conditions.append(
                and_(
                    Project.id.in_(db.session.query(robot_projects.c.id)),
                    Project.id.in_(db.session.query(system_projects.c.id))
                )
            )
        if "none" in project_type_filters:
            # Projetos que não aparecem em nenhuma das subqueries
            project_type_conditions.append(
                and_(
                    ~Project.id.in_(db.session.query(robot_projects.c.id)),
                    ~Project.id.in_(db.session.query(system_projects.c.id))
                )
            )
        
        # Se houver condições, aplica com OR (projeto pode atender qualquer uma)
        if project_type_conditions:
            if len(project_type_conditions) == 1:
                query = query.filter(project_type_conditions[0])
            else:
                query = query.filter(or_(*project_type_conditions))

    # Filtro por ferramentas/sistemas
    if tools_filter:
        tools_projects = db.session.query(Project.id).join(MacroStage).join(Stage).filter(
            or_(
                Stage.tools.like(f"%{tools_filter}%"),
                Stage.other_tools.like(f"%{tools_filter}%")
            )
        ).distinct().subquery()
        query = query.filter(Project.id.in_(db.session.query(tools_projects.c.id)))

    projects = query.order_by(Project.id).all()

    # Calcula status e progresso para cada projeto
    projects_with_status = []
    for project in projects:
        project_status = get_project_status(project)
        if project_status['is_manual']:
            progresso_percentual = None
        else:
            progresso_percentual = calculate_project_progress(project.start_date, project.end_date)
        
        projects_with_status.append({
            'project': project,
            'status': project_status['value'],
            'status_display': project_status['display_text'],
            'progress': progresso_percentual
        })
    
    # Aplica filtro de status (multi-select)
    if status_filters:
        projects_with_status = [
            item for item in projects_with_status 
            if item['status'] in status_filters
        ]
    
    # Aplica ordenação por progresso
    if progress_sort in ("asc", "desc"):
        def get_sort_key(item):
            progress = item['progress']
            if progress is None:
                return -1 if progress_sort == "desc" else 999
            return progress
        projects_with_status.sort(key=get_sort_key, reverse=(progress_sort == "desc"))

    # Cria Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Projetos"

    # Cabeçalhos
    headers = [
        "Nome do projeto",
        "Escopo",
        "Status",
        "Progresso (%)",
        "Link do GitHub",
        "Coordenador",
        "Equipe Automatiza / Suporte Automatiza",
        "Órgão demandante",
        "Setor interno",
        "Gestor responsável",
        "Contato do gestor responsável",
        "Gestor técnico",
        "Contato do gestor técnico",
        "Data de início",
        "Data de fim"
    ]
    
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    
    format_excel_header(ws, 1, len(headers))

    # Dados
    for row_idx, item in enumerate(projects_with_status, start=2):
        project = item['project']
        ws.cell(row=row_idx, column=1, value=project.name)
        ws.cell(row=row_idx, column=2, value=project.scope or "—")
        ws.cell(row=row_idx, column=3, value=item['status_display'] or "—")
        
        # Progresso
        if item['progress'] is not None:
            ws.cell(row=row_idx, column=4, value=f"{item['progress']}%")
        else:
            ws.cell(row=row_idx, column=4, value="—")
        
        ws.cell(row=row_idx, column=5, value=project.github_link or "—")
        ws.cell(row=row_idx, column=6, value=project.coordinator or "—")
        ws.cell(row=row_idx, column=7, value=project.automation_support or "—")
        ws.cell(row=row_idx, column=8, value=project.requesting_agency or "—")
        ws.cell(row=row_idx, column=9, value=project.internal_department or "—")
        ws.cell(row=row_idx, column=10, value=project.sponsoring_manager or "—")
        ws.cell(row=row_idx, column=11, value=project.sponsoring_manager_contact or "—")
        ws.cell(row=row_idx, column=12, value=project.technical_manager or "—")
        ws.cell(row=row_idx, column=13, value=project.technical_manager_contact or "—")
        
        # Datas
        format_excel_date(ws.cell(row=row_idx, column=14), project.start_date)
        format_excel_date(ws.cell(row=row_idx, column=15), project.end_date)

    # Ajusta larguras das colunas
    auto_adjust_column_width(ws)

    # Gera nome do arquivo com timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"projetos_{timestamp}"

    return create_excel_response(wb, filename)


@app.route("/dashboard/timeline", methods=["GET"])
def dashboard_timeline():
    """
    Dashboard de cronograma (Gantt) mostrando todos os projetos e suas hierarquias
    (macroetapas, etapas, tarefas) em uma linha do tempo visual.
    """
    # Lê parâmetros de filtro da query string (suporta múltiplos valores)
    project_name_filters = [p.strip() for p in request.args.getlist("project_name") if p.strip()]
    status_filters = [s.strip() for s in request.args.getlist("status") if s.strip()]
    requesting_agency_filters = [a.strip() for a in request.args.getlist("requesting_agency") if a.strip()]
    coordinator_filters = [c.strip() for c in request.args.getlist("coordinator") if c.strip()]
    project_type_filters = [p.strip() for p in request.args.getlist("project_type") if p.strip()]
    # Compatibilidade: se não houver lista, tenta valor único (ignora "all" que significa "todos")
    if not project_type_filters:
        single_project_type = request.args.get("project_type", "").strip()
        if single_project_type and single_project_type != "all":
            project_type_filters = [single_project_type]
    date_start_str = request.args.get("date_start", "").strip() or None
    date_end_str = request.args.get("date_end", "").strip() or None

    # Converte strings de data para objetos date
    date_start = None
    date_end = None
    if date_start_str:
        try:
            date_start = datetime.strptime(date_start_str, "%Y-%m-%d").date()
        except ValueError:
            date_start = None
    if date_end_str:
        try:
            date_end = datetime.strptime(date_end_str, "%Y-%m-%d").date()
        except ValueError:
            date_end = None

    # Inicia query base com eager loading para otimizar
    query = Project.query.options(
        joinedload(Project.macrostages).joinedload(MacroStage.stages).joinedload(Stage.tasks),
        joinedload(Project.macrostages).joinedload(MacroStage.tasks)
    )

    # Aplica filtros diretos em Project (multi-select)
    # Nota: filtro de status será aplicado após calcular status efetivo
    if project_name_filters:
        # Converte strings para inteiros (IDs de projeto)
        try:
            project_ids = [int(p_id) for p_id in project_name_filters]
            query = query.filter(Project.id.in_(project_ids))
        except ValueError:
            # Se houver erro na conversão, ignora o filtro
            pass
    if requesting_agency_filters:
        query = query.filter(Project.requesting_agency.in_(requesting_agency_filters))
    if coordinator_filters:
        query = query.filter(Project.coordinator.in_(coordinator_filters))

    # Filtro por tipo de projeto (robô/sistema) - multi-select
    if project_type_filters:
        robot_projects = db.session.query(Project.id).join(MacroStage).join(Stage).filter(
            Stage.stage_type == "robô"
        ).distinct().subquery()

        system_projects = db.session.query(Project.id).join(MacroStage).join(Stage).filter(
            Stage.stage_type == "sistema"
        ).distinct().subquery()

        # Lógica para múltiplos tipos selecionados
        project_type_conditions = []
        
        if "robot" in project_type_filters:
            project_type_conditions.append(Project.id.in_(db.session.query(robot_projects.c.id)))
        if "system" in project_type_filters:
            project_type_conditions.append(Project.id.in_(db.session.query(system_projects.c.id)))
        if "both" in project_type_filters:
            # Projetos que aparecem em ambas as subqueries
            project_type_conditions.append(
                and_(
                    Project.id.in_(db.session.query(robot_projects.c.id)),
                    Project.id.in_(db.session.query(system_projects.c.id))
                )
            )
        if "none" in project_type_filters:
            # Projetos que não aparecem em nenhuma das subqueries
            project_type_conditions.append(
                and_(
                    ~Project.id.in_(db.session.query(robot_projects.c.id)),
                    ~Project.id.in_(db.session.query(system_projects.c.id))
                )
            )
        
        # Se houver condições, aplica com OR (projeto pode atender qualquer uma)
        if project_type_conditions:
            if len(project_type_conditions) == 1:
                query = query.filter(project_type_conditions[0])
            else:
                query = query.filter(or_(*project_type_conditions))

    # Filtro por período
    if date_start or date_end:
        if date_start and date_end:
            # Projetos que iniciaram e terminaram dentro do período: start >= date_start AND end <= date_end
            query = query.filter(
                and_(
                    Project.start_date.isnot(None),
                    Project.end_date.isnot(None),
                    Project.start_date >= date_start,
                    Project.end_date <= date_end
                )
            )
        elif date_start:
            # Projetos que iniciaram a partir da data de início
            query = query.filter(
                and_(
                    Project.start_date.isnot(None),
                    Project.start_date >= date_start
                )
            )
        elif date_end:
            # Projetos que finalizam até a data de fim
            query = query.filter(
                and_(
                    Project.end_date.isnot(None),
                    Project.end_date <= date_end
                )
            )

    # Ordena por data de início ou nome
    projects = query.order_by(Project.start_date, Project.name).all()

    # Calcula status efetivo para cada projeto e aplica filtro de status
    # Nota: filtro de status deve ser aplicado após calcular status efetivo,
    # pois o status exibido é calculado dinamicamente, não vem do campo do banco
    projects_with_status = []
    for project in projects:
        project_status = get_project_status(project)
        projects_with_status.append({
            'project': project,
            'status': project_status['value']
        })
    
    # Aplica filtro de status após calcular status efetivo (multi-select)
    if status_filters:
        projects_with_status = [
            item for item in projects_with_status 
            if item['status'] in status_filters
        ]

    # Constrói estrutura hierárquica para o Gantt
    timeline_data = []

    for item in projects_with_status:
        project = item['project']
        # Só adiciona projeto se tiver datas válidas
        if not project.start_date or not project.end_date:
            continue

        project_item = {
            "id": f"project-{project.id}",
            "name": project.name,
            "start": project.start_date.isoformat(),
            "end": project.end_date.isoformat(),
            "progress": 0,
            "dependencies": None,  # frappe-gantt espera None ou array vazio para sem dependências
            "custom_class": "gantt-project",
            "url": url_for('project_detail', project_id=project.id),
        }
        timeline_data.append(project_item)

        # Adiciona macroetapas
        for macro in project.macrostages:
            if not macro.start_date or not macro.end_date:
                continue

            macro_item = {
                "id": f"macro-{macro.id}",
                "name": f"  {macro.name}",  # Indentação visual
                "start": macro.start_date.isoformat(),
                "end": macro.end_date.isoformat(),
                "progress": 0,
                "dependencies": f"project-{project.id}",
                "custom_class": "gantt-macro",
                "url": url_for('project_detail', project_id=project.id) + f"#macrostage-{macro.id}",
            }
            timeline_data.append(macro_item)

            # Adiciona etapas (se a macroetapa tiver estrutura_type="stages")
            if macro.structure_type == "stages":
                for stage in macro.stages:
                    if not stage.start_date or not stage.end_date:
                        continue

                    stage_item = {
                        "id": f"stage-{stage.id}",
                        "name": f"    {stage.name}",  # Mais indentado
                        "start": stage.start_date.isoformat(),
                        "end": stage.end_date.isoformat(),
                        "progress": 0,
                        "dependencies": f"macro-{macro.id}",
                        "custom_class": "gantt-stage",
                        "url": url_for('project_detail', project_id=project.id) + f"#stage-{stage.id}",
                    }
                    timeline_data.append(stage_item)

                    # Adiciona tarefas da etapa
                    for task in stage.tasks:
                        if not task.start_date or not task.end_date:
                            continue

                        task_item = {
                            "id": f"task-{task.id}",
                            "name": f"      {task.name}",  # Ainda mais indentado
                            "start": task.start_date.isoformat(),
                            "end": task.end_date.isoformat(),
                            "progress": 0,
                            "dependencies": f"stage-{stage.id}",
                            "custom_class": "gantt-task",
                            "url": url_for('project_detail', project_id=project.id) + f"#task-{task.id}",
                        }
                        timeline_data.append(task_item)

            # Adiciona tarefas diretas da macroetapa (se structure_type="tasks")
            elif macro.structure_type == "tasks":
                for task in macro.tasks:
                    if task.stage_id is not None:  # Ignora tarefas que pertencem a etapas
                        continue
                    if not task.start_date or not task.end_date:
                        continue

                    task_item = {
                        "id": f"task-{task.id}",
                        "name": f"    {task.name}",
                        "start": task.start_date.isoformat(),
                        "end": task.end_date.isoformat(),
                        "progress": 0,
                        "dependencies": f"macro-{macro.id}",
                        "custom_class": "gantt-task",
                        "url": url_for('project_detail', project_id=project.id) + f"#task-{task.id}",
                    }
                    timeline_data.append(task_item)

    # Coleta valores distintos para popular os selects dos filtros
    # Para status, calcula o status efetivo de todos os projetos para ter a lista completa
    # (não usa o campo status do banco, pois pode estar desatualizado)
    all_projects_for_status = Project.query.all()
    calculated_statuses = set()
    for p in all_projects_for_status:
        status_info = get_project_status(p)
        calculated_statuses.add(status_info['value'])
    
    filter_options = {
        "projects": sorted([
            (p.id, p.name) for p in Project.query.order_by(Project.name).all()
        ], key=lambda x: x[1]),  # Ordena por nome
        "statuses": sorted(list(calculated_statuses)),
        "requesting_agencies": sorted([
            a[0] for a in Project.query.with_entities(Project.requesting_agency).distinct()
            .filter(Project.requesting_agency.isnot(None)).all()
        ]),
        "coordinators": sorted([
            c[0] for c in Project.query.with_entities(Project.coordinator).distinct()
            .filter(Project.coordinator.isnot(None)).all()
        ]),
    }

    # Passa os valores atuais dos filtros para manter selecionados no template
    current_filters = {
        "project_name": project_name_filters,
        "status": status_filters,
        "requesting_agency": requesting_agency_filters,
        "coordinator": coordinator_filters,
        "project_type": project_type_filters,
        "date_start": date_start_str or "",
        "date_end": date_end_str or "",
    }

    return render_template(
        "dashboard_timeline.html",
        timeline_data=timeline_data,
        filter_options=filter_options,
        current_filters=current_filters,
    )


@app.route("/dashboard/timeline/export", methods=["GET"])
def export_dashboard_timeline():
    """
    Exporta o dashboard de cronograma para Excel (.xlsx).
    Reutiliza toda a lógica de filtros de dashboard_timeline().
    """
    # Lê parâmetros de filtro da query string (suporta múltiplos valores)
    project_name_filters = [p.strip() for p in request.args.getlist("project_name") if p.strip()]
    status_filters = [s.strip() for s in request.args.getlist("status") if s.strip()]
    requesting_agency_filters = [a.strip() for a in request.args.getlist("requesting_agency") if a.strip()]
    coordinator_filters = [c.strip() for c in request.args.getlist("coordinator") if c.strip()]
    project_type_filters = [p.strip() for p in request.args.getlist("project_type") if p.strip()]
    # Compatibilidade: se não houver lista, tenta valor único (ignora "all" que significa "todos")
    if not project_type_filters:
        single_project_type = request.args.get("project_type", "").strip()
        if single_project_type and single_project_type != "all":
            project_type_filters = [single_project_type]
    date_start_str = request.args.get("date_start", "").strip() or None
    date_end_str = request.args.get("date_end", "").strip() or None

    # Converte strings de data para objetos date
    date_start = None
    date_end = None
    if date_start_str:
        try:
            date_start = datetime.strptime(date_start_str, "%Y-%m-%d").date()
        except ValueError:
            pass
    if date_end_str:
        try:
            date_end = datetime.strptime(date_end_str, "%Y-%m-%d").date()
        except ValueError:
            pass

    # Reutiliza lógica de filtros do dashboard_timeline
    query = Project.query

    if project_name_filters:
        # Converte strings para inteiros (IDs de projeto)
        try:
            project_ids = [int(p_id) for p_id in project_name_filters]
            query = query.filter(Project.id.in_(project_ids))
        except ValueError:
            # Se houver erro na conversão, ignora o filtro
            pass
    if requesting_agency_filters:
        query = query.filter(Project.requesting_agency.in_(requesting_agency_filters))
    if coordinator_filters:
        query = query.filter(Project.coordinator.in_(coordinator_filters))

    # Filtro por tipo de projeto (robô/sistema) - multi-select
    if project_type_filters:
        robot_projects = db.session.query(Project.id).join(MacroStage).join(Stage).filter(
            Stage.stage_type == "robô"
        ).distinct().subquery()
        system_projects = db.session.query(Project.id).join(MacroStage).join(Stage).filter(
            Stage.stage_type == "sistema"
        ).distinct().subquery()

        # Lógica para múltiplos tipos selecionados
        project_type_conditions = []
        
        if "robot" in project_type_filters:
            project_type_conditions.append(Project.id.in_(db.session.query(robot_projects.c.id)))
        if "system" in project_type_filters:
            project_type_conditions.append(Project.id.in_(db.session.query(system_projects.c.id)))
        if "both" in project_type_filters:
            # Projetos que aparecem em ambas as subqueries
            project_type_conditions.append(
                and_(
                    Project.id.in_(db.session.query(robot_projects.c.id)),
                    Project.id.in_(db.session.query(system_projects.c.id))
                )
            )
        if "none" in project_type_filters:
            # Projetos que não aparecem em nenhuma das subqueries
            project_type_conditions.append(
                and_(
                    ~Project.id.in_(db.session.query(robot_projects.c.id)),
                    ~Project.id.in_(db.session.query(system_projects.c.id))
                )
            )
        
        # Se houver condições, aplica com OR (projeto pode atender qualquer uma)
        if project_type_conditions:
            if len(project_type_conditions) == 1:
                query = query.filter(project_type_conditions[0])
            else:
                query = query.filter(or_(*project_type_conditions))

    projects = query.options(
        joinedload(Project.macrostages).joinedload(MacroStage.stages).joinedload(Stage.tasks),
        joinedload(Project.macrostages).joinedload(MacroStage.tasks)
    ).order_by(Project.id).all()

    # Processa dados para exportação (similar ao dashboard_timeline)
    timeline_data = []
    for project in projects:
        project_status = get_project_status(project)
        
        # Aplica filtro de status após calcular (multi-select)
        if status_filters and project_status['value'] not in status_filters:
            continue
        
        # Adiciona projeto
        if project.start_date and project.end_date:
            # Verifica filtro de data
            if date_start and date_end:
                # Projetos que iniciaram e terminaram dentro do período
                if project.start_date < date_start or project.end_date > date_end:
                    continue
            elif date_start:
                # Projetos que iniciaram a partir da data de início
                if project.start_date < date_start:
                    continue
            elif date_end:
                # Projetos que finalizam até a data de fim
                if project.end_date > date_end:
                    continue
            
            project_item = {
                "name": project.name,
                "start": project.start_date,
                "end": project.end_date,
                "type": "Projeto"
            }
            timeline_data.append(project_item)

        # Processa macroetapas
        for macro in project.macrostages:
            if not macro.start_date or not macro.end_date:
                continue
            
            # Verifica filtro de data
            if date_start and date_end:
                # Macroetapas que iniciaram e terminaram dentro do período
                if macro.start_date < date_start or macro.end_date > date_end:
                    continue
            elif date_start:
                # Macroetapas que iniciaram a partir da data de início
                if macro.start_date < date_start:
                    continue
            elif date_end:
                # Macroetapas que finalizam até a data de fim
                if macro.end_date > date_end:
                    continue

            macro_item = {
                "name": f"  {macro.name}",
                "start": macro.start_date,
                "end": macro.end_date,
                "type": "Macroetapa",
                "project": project.name
            }
            timeline_data.append(macro_item)

            # Processa etapas
            if macro.structure_type == "stages":
                for stage in macro.stages:
                    if not stage.start_date or not stage.end_date:
                        continue
                    
                    # Verifica filtro de data
                    if date_start and date_end:
                        # Etapas que iniciaram e terminaram dentro do período
                        if stage.start_date < date_start or stage.end_date > date_end:
                            continue
                    elif date_start:
                        # Etapas que iniciaram a partir da data de início
                        if stage.start_date < date_start:
                            continue
                    elif date_end:
                        # Etapas que finalizam até a data de fim
                        if stage.end_date > date_end:
                            continue

                    stage_item = {
                        "name": f"    {stage.name}",
                        "start": stage.start_date,
                        "end": stage.end_date,
                        "type": "Etapa",
                        "project": project.name,
                        "macro": macro.name
                    }
                    timeline_data.append(stage_item)

                    # Processa tarefas da etapa
                    for task in stage.tasks:
                        if not task.start_date or not task.end_date:
                            continue
                        
                        # Verifica filtro de data
                        if date_start and date_end:
                            # Tarefas que iniciaram e terminaram dentro do período
                            if task.start_date < date_start or task.end_date > date_end:
                                continue
                        elif date_start:
                            # Tarefas que iniciaram a partir da data de início
                            if task.start_date < date_start:
                                continue
                        elif date_end:
                            # Tarefas que finalizam até a data de fim
                            if task.end_date > date_end:
                                continue

                        task_item = {
                            "name": f"      {task.name}",
                            "start": task.start_date,
                            "end": task.end_date,
                            "type": "Tarefa",
                            "project": project.name,
                            "macro": macro.name,
                            "stage": stage.name
                        }
                        timeline_data.append(task_item)

            # Processa tarefas diretas da macroetapa
            elif macro.structure_type == "tasks":
                for task in macro.tasks:
                    if task.stage_id is not None:
                        continue
                    if not task.start_date or not task.end_date:
                        continue
                    
                    # Verifica filtro de data
                    if date_start and date_end:
                        # Tarefas que iniciaram e terminaram dentro do período
                        if task.start_date < date_start or task.end_date > date_end:
                            continue
                    elif date_start:
                        # Tarefas que iniciaram a partir da data de início
                        if task.start_date < date_start:
                            continue
                    elif date_end:
                        # Tarefas que finalizam até a data de fim
                        if task.end_date > date_end:
                            continue

                    task_item = {
                        "name": f"    {task.name}",
                        "start": task.start_date,
                        "end": task.end_date,
                        "type": "Tarefa",
                        "project": project.name,
                        "macro": macro.name
                    }
                    timeline_data.append(task_item)

    # Cria Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Cronograma"

    # Cabeçalhos
    headers = [
        "Projeto",
        "Macroetapa",
        "Etapa",
        "Tarefa",
        "Tipo",
        "Data de início",
        "Data de fim",
        "Duração (dias)"
    ]
    
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    
    format_excel_header(ws, 1, len(headers))

    # Dados
    for row_idx, item in enumerate(timeline_data, start=2):
        ws.cell(row=row_idx, column=1, value=item.get('project', ''))
        ws.cell(row=row_idx, column=2, value=item.get('macro', ''))
        ws.cell(row=row_idx, column=3, value=item.get('stage', ''))
        
        # Nome (pode ser tarefa ou outro tipo)
        name_value = item.get('name', '').strip()
        if item.get('type') == 'Tarefa':
            ws.cell(row=row_idx, column=4, value=name_value)
        else:
            ws.cell(row=row_idx, column=4, value='')
        
        ws.cell(row=row_idx, column=5, value=item.get('type', ''))
        
        # Datas
        format_excel_date(ws.cell(row=row_idx, column=6), item.get('start'))
        format_excel_date(ws.cell(row=row_idx, column=7), item.get('end'))
        
        # Duração em dias
        if item.get('start') and item.get('end'):
            duration = (item['end'] - item['start']).days + 1
            ws.cell(row=row_idx, column=8, value=duration)
        else:
            ws.cell(row=row_idx, column=8, value="—")

    # Ajusta larguras das colunas
    auto_adjust_column_width(ws)

    # Gera nome do arquivo com timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"cronograma_{timestamp}"

    return create_excel_response(wb, filename)


@app.route("/dashboard/robots-systems", methods=["GET"])
def dashboard_robots_systems():
    """
    Dashboard de Robôs e Sistemas.
    Lista todas as etapas do tipo "robô" ou "sistema" de todos os projetos,
    com cálculo de status baseado nas tarefas associadas.
    """
    # Lê parâmetros de filtro da query string (suporta múltiplos valores)
    stage_type_filters = [s.strip() for s in request.args.getlist("stage_type_filter") if s.strip()]
    # Compatibilidade: se não houver lista, tenta valor único (ignora "todos" que significa "todos")
    if not stage_type_filters:
        single_stage_type = request.args.get("stage_type_filter", "").strip()
        if single_stage_type and single_stage_type != "todos":
            stage_type_filters = [single_stage_type]
    
    status_filters = [s.strip() for s in request.args.getlist("status_filter") if s.strip()]
    # Compatibilidade: se não houver lista, tenta valor único (ignora "todos" que significa "todos")
    if not status_filters:
        single_status = request.args.get("status_filter", "").strip()
        if single_status and single_status != "todos":
            status_filters = [single_status]
    
    project_id_filters = [p.strip() for p in request.args.getlist("project_id") if p.strip()]
    requesting_agency_filters = [a.strip() for a in request.args.getlist("requesting_agency") if a.strip()]
    
    # Inicia query base: busca todas as etapas do tipo robô ou sistema
    query = Stage.query.join(MacroStage).join(Project).filter(
        Stage.stage_type.in_(["robô", "sistema"])
    )
    
    # Aplica filtro de tipo de etapa (multi-select)
    if stage_type_filters:
        query = query.filter(Stage.stage_type.in_(stage_type_filters))
    
    # Aplica filtro de projeto (multi-select)
    if project_id_filters:
        try:
            project_id_ints = [int(p) for p in project_id_filters]
            query = query.filter(Project.id.in_(project_id_ints))
        except ValueError:
            pass  # Ignora se não for um número válido
    
    # Aplica filtro de órgão demandante (multi-select)
    if requesting_agency_filters:
        query = query.filter(Project.requesting_agency.in_(requesting_agency_filters))
    
    # Carrega relacionamentos necessários
    stages = query.options(
        joinedload(Stage.macrostage).joinedload(MacroStage.project),
        joinedload(Stage.tasks)
    ).order_by(Project.name, MacroStage.name, Stage.name).all()
    
    # Processa cada etapa: calcula datas agregadas e status
    robots_and_systems = []
    for stage in stages:
        # Calcula datas agregadas das tarefas
        tasks = stage.tasks
        start_dates = [t.start_date for t in tasks if t.start_date is not None]
        end_dates = [t.end_date for t in tasks if t.end_date is not None]
        
        etapa_start_date = min(start_dates) if start_dates else None
        etapa_end_date = max(end_dates) if end_dates else None
        
        # Calcula status baseado nas tarefas
        calculated_status = calculate_stage_status(stage)
        
        # Aplica filtro de status (multi-select)
        if status_filters:
            if calculated_status not in status_filters:
                continue
        
        # Monta estrutura de dados
        robots_and_systems.append({
            "id": stage.id,
            "stage_name": stage.name,
            "stage_type": stage.stage_type,
            "project_id": stage.macrostage.project.id,
            "project_name": stage.macrostage.project.name,
            "macrostage_name": stage.macrostage.name,
            "scope": stage.scope,
            "tools": stage.tools,
            "other_tools": stage.other_tools,
            "start_date": etapa_start_date,
            "end_date": etapa_end_date,
            "status": calculated_status,
        })
    
    # Coleta valores distintos para popular os selects dos filtros
    # Projetos que possuem robôs/sistemas
    projects_with_robots = db.session.query(Project.id, Project.name).join(
        MacroStage
    ).join(Stage).filter(
        Stage.stage_type.in_(["robô", "sistema"])
    ).distinct().order_by(Project.name).all()
    
    # Órgãos demandantes dos projetos que possuem robôs/sistemas
    requesting_agencies = db.session.query(Project.requesting_agency).join(
        MacroStage
    ).join(Stage).filter(
        Stage.stage_type.in_(["robô", "sistema"]),
        Project.requesting_agency.isnot(None)
    ).distinct().order_by(Project.requesting_agency).all()
    requesting_agencies = [a[0] for a in requesting_agencies]
    
    # Passa os valores atuais dos filtros para manter selecionados no template (como listas)
    current_filters = {
        "stage_type": stage_type_filters,
        "status": status_filters,
        "project_id": project_id_filters,
        "requesting_agency": requesting_agency_filters,
    }
    
    return render_template(
        "dashboard_robots_systems.html",
        robots_and_systems=robots_and_systems,
        filter_options={
            "projects": projects_with_robots,
            "requesting_agencies": requesting_agencies,
        },
        current_filters=current_filters,
    )


@app.route("/dashboard/robots-systems/export", methods=["GET"])
def export_dashboard_robots_systems():
    """
    Exporta o dashboard de robôs e sistemas para Excel (.xlsx).
    Reutiliza toda a lógica de filtros de dashboard_robots_systems().
    """
    # Lê parâmetros de filtro da query string (suporta múltiplos valores)
    stage_type_filters = [s.strip() for s in request.args.getlist("stage_type_filter") if s.strip()]
    # Compatibilidade: se não houver lista, tenta valor único (ignora "todos" que significa "todos")
    if not stage_type_filters:
        single_stage_type = request.args.get("stage_type_filter", "").strip()
        if single_stage_type and single_stage_type != "todos":
            stage_type_filters = [single_stage_type]
    
    status_filters = [s.strip() for s in request.args.getlist("status_filter") if s.strip()]
    # Compatibilidade: se não houver lista, tenta valor único (ignora "todos" que significa "todos")
    if not status_filters:
        single_status = request.args.get("status_filter", "").strip()
        if single_status and single_status != "todos":
            status_filters = [single_status]
    
    project_id_filters = [p.strip() for p in request.args.getlist("project_id") if p.strip()]
    requesting_agency_filters = [a.strip() for a in request.args.getlist("requesting_agency") if a.strip()]
    
    # Inicia query base
    query = Stage.query.join(MacroStage).join(Project).filter(
        Stage.stage_type.in_(["robô", "sistema"])
    )
    
    # Aplica filtros (multi-select)
    if stage_type_filters and "todos" not in stage_type_filters:
        query = query.filter(Stage.stage_type.in_(stage_type_filters))
    
    if project_id_filters:
        try:
            project_id_ints = [int(p) for p in project_id_filters]
            query = query.filter(Project.id.in_(project_id_ints))
        except ValueError:
            pass
    
    if requesting_agency_filters:
        query = query.filter(Project.requesting_agency.in_(requesting_agency_filters))
    
    stages = query.options(
        joinedload(Stage.macrostage).joinedload(MacroStage.project),
        joinedload(Stage.tasks)
    ).order_by(Project.name, MacroStage.name, Stage.name).all()
    
    # Processa cada etapa
    robots_and_systems = []
    for stage in stages:
        tasks = stage.tasks
        start_dates = [t.start_date for t in tasks if t.start_date is not None]
        end_dates = [t.end_date for t in tasks if t.end_date is not None]
        
        etapa_start_date = min(start_dates) if start_dates else None
        etapa_end_date = max(end_dates) if end_dates else None
        
        calculated_status = calculate_stage_status(stage)
        
        # Aplica filtro de status (multi-select)
        if status_filters:
            if calculated_status not in status_filters:
                continue
        
        robots_and_systems.append({
            "id": stage.id,
            "stage_name": stage.name,
            "stage_type": stage.stage_type,
            "project_id": stage.macrostage.project.id,
            "project_name": stage.macrostage.project.name,
            "macrostage_name": stage.macrostage.name,
            "scope": stage.scope,
            "tools": stage.tools,
            "other_tools": stage.other_tools,
            "start_date": etapa_start_date,
            "end_date": etapa_end_date,
            "status": calculated_status,
        })

    # Cria Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Robôs e Sistemas"

    # Cabeçalhos
    headers = [
        "Nome do robô/sistema",
        "Tipo",
        "Projeto",
        "Macroetapa",
        "Escopo",
        "Ferramentas",
        "Data de início",
        "Data de fim",
        "Status"
    ]
    
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    
    format_excel_header(ws, 1, len(headers))

    # Dados
    for row_idx, item in enumerate(robots_and_systems, start=2):
        ws.cell(row=row_idx, column=1, value=item['stage_name'])
        ws.cell(row=row_idx, column=2, value=item['stage_type'])
        ws.cell(row=row_idx, column=3, value=item['project_name'])
        ws.cell(row=row_idx, column=4, value=item['macrostage_name'])
        ws.cell(row=row_idx, column=5, value=item['scope'] or "—")
        
        # Ferramentas (combina tools e other_tools)
        tools_list = []
        if item['tools']:
            tools_list.append(item['tools'])
        if item['other_tools']:
            tools_list.append(item['other_tools'])
        ws.cell(row=row_idx, column=6, value=", ".join(tools_list) if tools_list else "—")
        
        # Datas
        format_excel_date(ws.cell(row=row_idx, column=7), item['start_date'])
        format_excel_date(ws.cell(row=row_idx, column=8), item['end_date'])
        
        ws.cell(row=row_idx, column=9, value=item['status'])

    # Ajusta larguras das colunas
    auto_adjust_column_width(ws)

    # Gera nome do arquivo com timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"robos_sistemas_{timestamp}"

    return create_excel_response(wb, filename)


def create_project_from_template(name: str) -> Project:
    """
    Cria um novo projeto a partir do template pré-definido.
    
    Args:
        name: Nome do projeto
        
    Returns:
        Project: Projeto criado com toda a estrutura de macroetapas, etapas e tarefas
    """
    # Criar projeto
    project = Project(name=name)
    db.session.add(project)
    db.session.flush()  # Para obter o ID do projeto
    
    # Data inicial: hoje
    current_date = date.today()
    
    # Macroetapa 1: "Alinhamentos iniciais com demandante" (sem etapas)
    macrostage1 = MacroStage(
        project=project,
        name="Alinhamentos iniciais com demandante",
        position=1,
        structure_type="tasks"
    )
    db.session.add(macrostage1)
    db.session.flush()
    
    # Tarefa 1 da Macroetapa 1: 7 dias
    task1_start = current_date
    task1_end = task1_start + timedelta(days=6)  # 7 dias incluindo início e fim
    task1 = Task(
        name="Reuniões de alinhamentos iniciais com o demandante e definição da Equipe Automatiza que irá atuar no Projeto",
        macrostage=macrostage1,
        stage=None,
        start_date=task1_start,
        end_date=task1_end,
        position=1
    )
    db.session.add(task1)
    current_date = task1_end + timedelta(days=1)  # Próxima tarefa começa no dia seguinte
    
    # Macroetapa 2: "Imersão para entendimento da demanda" (sem etapas)
    macrostage2 = MacroStage(
        project=project,
        name="Imersão para entendimento da demanda",
        position=2,
        structure_type="tasks"
    )
    db.session.add(macrostage2)
    db.session.flush()
    
    # Tarefa 1 da Macroetapa 2: 5 dias
    task2_start = current_date
    task2_end = task2_start + timedelta(days=4)  # 5 dias
    task2 = Task(
        name="Análise dos documentos e informações repassadas pelo demandante",
        macrostage=macrostage2,
        stage=None,
        start_date=task2_start,
        end_date=task2_end,
        position=1
    )
    db.session.add(task2)
    current_date = task2_end + timedelta(days=1)
    
    # Tarefa 2 da Macroetapa 2: 5 dias
    task3_start = current_date
    task3_end = task3_start + timedelta(days=4)  # 5 dias
    task3 = Task(
        name="Avaliação de quais ferramentas e tecnologias poderão ser utilizadas para solucionar o desafio",
        macrostage=macrostage2,
        stage=None,
        start_date=task3_start,
        end_date=task3_end,
        position=2
    )
    db.session.add(task3)
    current_date = task3_end + timedelta(days=1)
    
    # Macroetapa 3: "Desenvolvimento de ferramentas" (com etapas)
    macrostage3 = MacroStage(
        project=project,
        name="Desenvolvimento de ferramentas",
        position=3,
        structure_type="stages"
    )
    db.session.add(macrostage3)
    db.session.flush()
    
    # Etapa 1: "Desenvolvimento de robô"
    stage1 = Stage(
        name="Desenvolvimento de robô",
        macrostage=macrostage3,
        position=1,
        stage_type="robô"
    )
    db.session.add(stage1)
    db.session.flush()
    
    # Tarefas da Etapa 1 (robô)
    # Tarefa 1: 20 dias
    task4_start = current_date
    task4_end = task4_start + timedelta(days=19)  # 20 dias
    task4 = Task(
        name="Desenvolver e testar solução localmente",
        macrostage=macrostage3,
        stage=stage1,
        start_date=task4_start,
        end_date=task4_end,
        position=1
    )
    db.session.add(task4)
    current_date = task4_end + timedelta(days=1)
    
    # Tarefa 2: 10 dias
    task5_start = current_date
    task5_end = task5_start + timedelta(days=9)  # 10 dias
    task5 = Task(
        name="Testar solução com o demandante e realizar alterações necessárias",
        macrostage=macrostage3,
        stage=stage1,
        start_date=task5_start,
        end_date=task5_end,
        position=2
    )
    db.session.add(task5)
    current_date = task5_end + timedelta(days=1)
    
    # Tarefa 3: 10 dias
    task6_start = current_date
    task6_end = task6_start + timedelta(days=9)  # 10 dias
    task6 = Task(
        name="Implantar versão final da solução",
        macrostage=macrostage3,
        stage=stage1,
        start_date=task6_start,
        end_date=task6_end,
        position=3
    )
    db.session.add(task6)
    # Para a etapa de sistema, vamos começar após o término da etapa de robô
    system_start_date = task6_end + timedelta(days=1)
    
    # Etapa 2: "Desenvolvimento de sistema"
    stage2 = Stage(
        name="Desenvolvimento de sistema",
        macrostage=macrostage3,
        position=2,
        stage_type="sistema"
    )
    db.session.add(stage2)
    db.session.flush()
    
    # Tarefas da Etapa 2 (sistema) - começam após a última tarefa do robô
    # Tarefa 1: 20 dias
    task7_start = system_start_date
    task7_end = task7_start + timedelta(days=19)  # 20 dias
    task7 = Task(
        name="Desenvolver e testar solução localmente",
        macrostage=macrostage3,
        stage=stage2,
        start_date=task7_start,
        end_date=task7_end,
        position=1
    )
    db.session.add(task7)
    current_date = task7_end + timedelta(days=1)
    
    # Tarefa 2: 10 dias
    task8_start = current_date
    task8_end = task8_start + timedelta(days=9)  # 10 dias
    task8 = Task(
        name="Testar solução com o demandante e realizar alterações necessárias",
        macrostage=macrostage3,
        stage=stage2,
        start_date=task8_start,
        end_date=task8_end,
        position=2
    )
    db.session.add(task8)
    current_date = task8_end + timedelta(days=1)
    
    # Tarefa 3: 10 dias
    task9_start = current_date
    task9_end = task9_start + timedelta(days=9)  # 10 dias
    task9 = Task(
        name="Implantar versão final da solução",
        macrostage=macrostage3,
        stage=stage2,
        start_date=task9_start,
        end_date=task9_end,
        position=3
    )
    db.session.add(task9)
    current_date = task9_end + timedelta(days=1)
    
    # Macroetapa 4: "Encerramentos com demandante" (sem etapas)
    macrostage4 = MacroStage(
        project=project,
        name="Encerramentos com demandante",
        position=4,
        structure_type="tasks"
    )
    db.session.add(macrostage4)
    db.session.flush()
    
    # Tarefa da Macroetapa 4: 7 dias
    task10_start = current_date
    task10_end = task10_start + timedelta(days=6)  # 7 dias
    task10 = Task(
        name="Reuniões de alinhamentos finais com o demandante",
        macrostage=macrostage4,
        stage=None,
        start_date=task10_start,
        end_date=task10_end,
        position=1
    )
    db.session.add(task10)
    
    # Commit todas as alterações
    db.session.commit()
    
    # Recalcular todas as datas em cascata
    recalculate_stage(stage1)
    recalculate_stage(stage2)
    recalculate_macrostage(macrostage1)
    recalculate_macrostage(macrostage2)
    recalculate_macrostage(macrostage3)
    recalculate_macrostage(macrostage4)
    recalculate_project(project)
    recalculate_project_status(project)
    db.session.commit()
    
    return project


@app.route("/projects/create", methods=["POST"])
def create_project():
    """
    Cria um novo Projeto com base no nome enviado via formulário.
    As datas de início e fim serão calculadas depois, a partir das macroetapas.
    """
    name = request.form.get("name", "").strip()
    if not name:
        # Em uma versão mais elaborada, poderíamos exibir mensagem de erro.
        return redirect(url_for("list_projects"))
    
    project_type = request.form.get("project_type", "blank").strip()
    
    if project_type == "template":
        # Criar projeto a partir do template
        project = create_project_from_template(name)
    else:
        # Criar projeto em branco (comportamento original)
        project = Project(name=name)
        db.session.add(project)
        db.session.commit()
    
    return redirect(url_for("list_projects"))


@app.route("/projects/<int:project_id>", methods=["GET"])
def project_detail(project_id: int):
    """
    Exibe a página de detalhe de um Projeto, com suas macroetapas, etapas e tarefas.
    Permite adicionar Macroetapas, Etapas e Tarefas via formulários simples.
    """
    project = Project.query.get_or_404(project_id)
    # Calcula o status efetivo (manual ou automático)
    project_status = get_project_status(project)
    # Calcula o progresso temporal do projeto
    # Se o status for manual (Suspenso ou Descartado), não calcular progresso
    if project_status['is_manual']:
        progresso_percentual = None
    else:
        progresso_percentual = calculate_project_progress(project.start_date, project.end_date)
    # As macroetapas e etapas serão acessadas via relacionamentos no template.
    return render_template(
        "project_detail.html",
        project=project,
        status_choices=PROJECT_STATUS_CHOICES,
        stage_type_choices=STAGE_TYPE_CHOICES,
        progresso_percentual=progresso_percentual,
        project_status=project_status,
    )


@app.route("/projects/<int:project_id>/export", methods=["GET"])
def export_project_detail(project_id: int):
    """
    Exporta o relatório completo do projeto para Excel (.xlsx).
    Cria múltiplas planilhas com todas as informações do projeto.
    """
    project = Project.query.options(
        joinedload(Project.macrostages).joinedload(MacroStage.stages).joinedload(Stage.tasks),
        joinedload(Project.macrostages).joinedload(MacroStage.tasks)
    ).get_or_404(project_id)
    
    # Calcula status e progresso
    project_status = get_project_status(project)
    if project_status['is_manual']:
        progresso_percentual = None
    else:
        progresso_percentual = calculate_project_progress(project.start_date, project.end_date)
    
    # Cria Workbook
    wb = Workbook()
    
    # Remove planilha padrão
    if wb.active:
        wb.remove(wb.active)
    
    # ===== Planilha 1: Informações do Projeto =====
    ws_info = wb.create_sheet("Informações do Projeto")
    
    info_data = [
        ["Campo", "Valor"],
        ["Nome do projeto", project.name],
        ["Escopo", project.scope or "—"],
        ["Status", project_status['display_text'] or "—"],
        ["Progresso", f"{progresso_percentual}%" if progresso_percentual is not None else "—"],
        ["Data de início (calculada)", project.start_date.strftime("%d/%m/%Y") if project.start_date else "—"],
        ["Data de fim (calculada)", project.end_date.strftime("%d/%m/%Y") if project.end_date else "—"],
        ["Link do GitHub", project.github_link or "—"],
        ["Coordenador", project.coordinator or "—"],
        ["Equipe Automatiza / Suporte Automatiza", project.automation_support or "—"],
        ["Órgão demandante", project.requesting_agency or "—"],
        ["Setor interno", project.internal_department or "—"],
        ["Gestor responsável", project.sponsoring_manager or "—"],
        ["Contato do gestor responsável", project.sponsoring_manager_contact or "—"],
        ["Gestor técnico", project.technical_manager or "—"],
        ["Contato do gestor técnico", project.technical_manager_contact or "—"],
    ]
    
    for row_idx, row_data in enumerate(info_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_info.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:  # Cabeçalho
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    auto_adjust_column_width(ws_info)
    
    # ===== Planilha 2: Macroetapas =====
    ws_macros = wb.create_sheet("Macroetapas")
    
    headers_macros = ["Nome", "Tipo de estrutura", "Data de início", "Data de fim", "Posição"]
    for col, header in enumerate(headers_macros, start=1):
        ws_macros.cell(row=1, column=col, value=header)
    format_excel_header(ws_macros, 1, len(headers_macros))
    
    for row_idx, macro in enumerate(project.macrostages, start=2):
        ws_macros.cell(row=row_idx, column=1, value=macro.name)
        ws_macros.cell(row=row_idx, column=2, value=macro.structure_type or "—")
        format_excel_date(ws_macros.cell(row=row_idx, column=3), macro.start_date)
        format_excel_date(ws_macros.cell(row=row_idx, column=4), macro.end_date)
        ws_macros.cell(row=row_idx, column=5, value=macro.position)
    
    auto_adjust_column_width(ws_macros)
    
    # ===== Planilha 3: Etapas =====
    ws_stages = wb.create_sheet("Etapas")
    
    headers_stages = ["Macroetapa", "Nome", "Tipo", "Escopo", "Ferramentas", "Outras ferramentas", "Data de início", "Data de fim", "Posição"]
    for col, header in enumerate(headers_stages, start=1):
        ws_stages.cell(row=1, column=col, value=header)
    format_excel_header(ws_stages, 1, len(headers_stages))
    
    row_idx = 2
    for macro in project.macrostages:
        for stage in macro.stages:
            ws_stages.cell(row=row_idx, column=1, value=macro.name)
            ws_stages.cell(row=row_idx, column=2, value=stage.name)
            ws_stages.cell(row=row_idx, column=3, value=stage.stage_type or "—")
            ws_stages.cell(row=row_idx, column=4, value=stage.scope or "—")
            ws_stages.cell(row=row_idx, column=5, value=stage.tools or "—")
            ws_stages.cell(row=row_idx, column=6, value=stage.other_tools or "—")
            format_excel_date(ws_stages.cell(row=row_idx, column=7), stage.start_date)
            format_excel_date(ws_stages.cell(row=row_idx, column=8), stage.end_date)
            ws_stages.cell(row=row_idx, column=9, value=stage.position)
            row_idx += 1
    
    auto_adjust_column_width(ws_stages)
    
    # ===== Planilha 4: Tarefas =====
    ws_tasks = wb.create_sheet("Tarefas")
    
    headers_tasks = ["Macroetapa", "Etapa", "Nome", "Data de início", "Data de fim", "Posição"]
    for col, header in enumerate(headers_tasks, start=1):
        ws_tasks.cell(row=1, column=col, value=header)
    format_excel_header(ws_tasks, 1, len(headers_tasks))
    
    row_idx = 2
    for macro in project.macrostages:
        # Tarefas de etapas
        for stage in macro.stages:
            for task in stage.tasks:
                ws_tasks.cell(row=row_idx, column=1, value=macro.name)
                ws_tasks.cell(row=row_idx, column=2, value=stage.name)
                ws_tasks.cell(row=row_idx, column=3, value=task.name)
                format_excel_date(ws_tasks.cell(row=row_idx, column=4), task.start_date)
                format_excel_date(ws_tasks.cell(row=row_idx, column=5), task.end_date)
                ws_tasks.cell(row=row_idx, column=6, value=task.position)
                row_idx += 1
        
        # Tarefas diretas da macroetapa
        for task in macro.tasks:
            if task.stage_id is None:  # Apenas tarefas diretas
                ws_tasks.cell(row=row_idx, column=1, value=macro.name)
                ws_tasks.cell(row=row_idx, column=2, value="—")  # Sem etapa
                ws_tasks.cell(row=row_idx, column=3, value=task.name)
                format_excel_date(ws_tasks.cell(row=row_idx, column=4), task.start_date)
                format_excel_date(ws_tasks.cell(row=row_idx, column=5), task.end_date)
                ws_tasks.cell(row=row_idx, column=6, value=task.position)
                row_idx += 1
    
    auto_adjust_column_width(ws_tasks)
    
    # ===== Planilha 5: Atualizações Semanais =====
    # Coleta todas as atualizações semanais do projeto
    all_updates = []
    for macro in project.macrostages:
        # Tarefas diretas da macroetapa
        for task in macro.tasks:
            if task.stage_id is None:  # Apenas tarefas diretas
                for update in task.weekly_updates:
                    all_updates.append({
                        'task_name': task.name,
                        'macro_name': macro.name,
                        'stage_name': None,
                        'update_date': update.update_date,
                        'content': update.content
                    })
        # Tarefas de etapas
        for stage in macro.stages:
            for task in stage.tasks:
                for update in task.weekly_updates:
                    all_updates.append({
                        'task_name': task.name,
                        'macro_name': macro.name,
                        'stage_name': stage.name,
                        'update_date': update.update_date,
                        'content': update.content
                    })
    
    if all_updates:
        ws_updates = wb.create_sheet("Atualizações Semanais")
        
        headers_updates = ["Tarefa", "Macroetapa", "Etapa", "Data da atualização", "Conteúdo"]
        for col, header in enumerate(headers_updates, start=1):
            ws_updates.cell(row=1, column=col, value=header)
        format_excel_header(ws_updates, 1, len(headers_updates))
        
        for row_idx, update in enumerate(all_updates, start=2):
            ws_updates.cell(row=row_idx, column=1, value=update['task_name'])
            ws_updates.cell(row=row_idx, column=2, value=update['macro_name'])
            ws_updates.cell(row=row_idx, column=3, value=update['stage_name'] or "—")
            format_excel_date(ws_updates.cell(row=row_idx, column=4), update['update_date'])
            ws_updates.cell(row=row_idx, column=5, value=update['content'])
            # Permite quebra de linha no conteúdo
            ws_updates.cell(row=row_idx, column=5).alignment = Alignment(wrap_text=True)
        
        auto_adjust_column_width(ws_updates)
    
    # Gera nome do arquivo com timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    sanitized_name = sanitize_filename(project.name)
    filename = f"projeto_{sanitized_name}_{timestamp}"

    return create_excel_response(wb, filename)


@app.route("/macrostages/create", methods=["POST"])
def create_macrostage():
    """
    Cria uma nova Macroetapa ligada a um Projeto.
    """
    project_id = request.form.get("project_id")
    name = request.form.get("name", "").strip()

    if not project_id or not name:
        return redirect(url_for("list_projects"))

    project = Project.query.get(project_id)
    if project is None:
        abort(404)

    existing_positions = [m.position or 0 for m in project.macrostages]
    next_position = (max(existing_positions) + 1) if existing_positions else 1

    macrostage = MacroStage(name=name, project=project, position=next_position)
    db.session.add(macrostage)
    db.session.commit()

    # Recalcula datas do projeto (caso a macroetapa venha a ter datas no futuro)
    recalculate_project(project)
    recalculate_project_status(project)
    db.session.commit()

    return redirect_with_anchor("project_detail", f"macrostage-{macrostage.id}", project_id=project.id)


@app.route("/stages/create", methods=["POST"])
def create_stage():
    """
    Cria uma nova Etapa ligada a uma Macroetapa.
    """
    macrostage_id = request.form.get("macrostage_id")
    name = request.form.get("name", "").strip()
    stage_type = (request.form.get("stage_type", "") or "não se aplica").strip()
    scope = request.form.get("scope", "").strip()
    tools_selected = request.form.getlist("tools")
    other_tools = request.form.get("other_tools", "").strip()

    if stage_type not in STAGE_TYPE_CHOICES:
        stage_type = "não se aplica"

    if stage_type not in ("robô", "sistema"):
        scope = None
        tools_selected = []
        other_tools = None
    else:
        scope = scope or None
        other_tools = other_tools or None

    tools_str = ",".join(tools_selected) if tools_selected else None

    if not macrostage_id or not name:
        return redirect(url_for("list_projects"))

    macrostage = MacroStage.query.get(macrostage_id)
    if macrostage is None:
        abort(404)

    if macrostage.structure_type == "tasks":
        return redirect(url_for("project_detail", project_id=macrostage.project.id))

    if macrostage.structure_type is None:
        macrostage.structure_type = "stages"

    existing_positions = [s.position or 0 for s in macrostage.stages]
    next_position = (max(existing_positions) + 1) if existing_positions else 1

    stage = Stage(
        name=name,
        macrostage=macrostage,
        position=next_position,
        stage_type=stage_type,
        scope=scope,
        tools=tools_str,
        other_tools=other_tools,
    )
    db.session.add(stage)
    db.session.commit()

    # Recalcula datas da macroetapa/projeto (ainda não há tarefas, então provavelmente continuará None)
    recalculate_macrostage(macrostage)
    recalculate_project(macrostage.project)
    recalculate_project_status(macrostage.project)
    db.session.commit()

    return redirect_with_anchor("project_detail", f"stage-{stage.id}", project_id=macrostage.project.id)


@app.route("/tasks/create", methods=["POST"])
def create_task():
    """
    Cria uma nova Tarefa ligada a uma Etapa.

    Recebe do formulário:
        - stage_id
        - name
        - start_date (YYYY-MM-DD)
        - end_date   (YYYY-MM-DD)
    """
    stage_id = request.form.get("stage_id")
    macrostage_id = request.form.get("macrostage_id")
    name = request.form.get("name", "").strip()
    start_date_str = request.form.get("start_date", "").strip()
    end_date_str = request.form.get("end_date", "").strip()

    if not name:
        return redirect(url_for("list_projects"))

    stage = Stage.query.get(stage_id) if stage_id else None
    if stage_id and stage is None:
        abort(404)

    if stage:
        macrostage = stage.macrostage
    else:
        if not macrostage_id:
            return redirect(url_for("list_projects"))
        macrostage = MacroStage.query.get(macrostage_id)
        if macrostage is None:
            abort(404)

    # Conversão das strings de data (HTML <input type="date">) para objetos date
    start_date = parse_date_field(start_date_str)
    end_date = parse_date_field(end_date_str)

    # Validação das datas
    is_valid, error_message = validate_task_dates(start_date, end_date)
    if not is_valid:
        project_id = macrostage.project.id
        if stage:
            anchor = f"stage-{stage.id}"
        else:
            anchor = f"macrostage-{macrostage.id}"
        url = url_for("project_detail", project_id=project_id, error=quote(error_message))
        if anchor:
            url = f"{url}#{anchor}"
        return redirect(url)

    if stage:
        if stage.macrostage.structure_type == "tasks":
            return redirect(url_for("project_detail", project_id=stage.macrostage.project.id))
        existing_positions = [t.position or 0 for t in stage.tasks]
    else:
        if macrostage.structure_type == "stages":
            return redirect(url_for("project_detail", project_id=macrostage.project.id))
        if macrostage.structure_type is None:
            macrostage.structure_type = "tasks"
        existing_positions = [t.position or 0 for t in macrostage.tasks if t.stage is None]

    next_position = (max(existing_positions) + 1) if existing_positions else 1

    task = Task(
        name=name,
        stage=stage,
        macrostage=macrostage,
        start_date=start_date,
        end_date=end_date,
        position=next_position,
    )
    db.session.add(task)
    
    # Tenta fazer commit, capturando exceções de validação do modelo
    try:
        db.session.commit()
    except ValueError as e:
        db.session.rollback()
        project_id = macrostage.project.id
        if stage:
            anchor = f"stage-{stage.id}"
        else:
            anchor = f"macrostage-{macrostage.id}"
        url = url_for("project_detail", project_id=project_id, error=quote(str(e)))
        if anchor:
            url = f"{url}#{anchor}"
        return redirect(url)

    # Recalcula datas em cascata (etapa, macroetapa, projeto)
    if stage:
        recalculate_all_from_stage(stage)
        anchor = f"stage-{stage.id}"
    else:
        recalculate_all_from_macrostage(macrostage)
        anchor = f"macrostage-{macrostage.id}"

    return redirect_with_anchor("project_detail", anchor, project_id=macrostage.project.id)


@app.route("/projects/<int:project_id>/update", methods=["POST"])
def update_project(project_id: int):
    """
    Atualiza as informações gerais de um projeto.
    """
    project = Project.query.get_or_404(project_id)
    name = request.form.get("name", "").strip()

    if name:
        project.name = name

    project.scope = request.form.get("scope", "").strip() or None
    status = request.form.get("status", "").strip()
    
    # Tratamento de status manual vs automático
    if status in ("Suspenso", "Descartado"):
        # Status manual
        project.status_manual = True
        project.status_manual_value = status
        project.status = status
    elif status in PROJECT_STATUS_CHOICES:
        # Outro status válido - volta para automático
        project.status_manual = False
        project.status_manual_value = None
        # Calcula status automático
        automatic_status = calculate_automatic_status(project)
        project.status = automatic_status
    else:
        # Status vazio ou inválido - volta para automático
        project.status_manual = False
        project.status_manual_value = None
        # Calcula status automático
        automatic_status = calculate_automatic_status(project)
        project.status = automatic_status
    
    project.github_link = request.form.get("github_link", "").strip() or None
    project.coordinator = request.form.get("coordinator", "").strip() or None
    project.automation_support = request.form.get("automation_support", "").strip() or None
    project.requesting_agency = request.form.get("requesting_agency", "").strip() or None
    project.internal_department = request.form.get("internal_department", "").strip() or None
    project.sponsoring_manager = request.form.get("sponsoring_manager", "").strip() or None
    project.sponsoring_manager_contact = request.form.get("sponsoring_manager_contact", "").strip() or None
    project.technical_manager = request.form.get("technical_manager", "").strip() or None
    project.technical_manager_contact = request.form.get("technical_manager_contact", "").strip() or None
    
    # Processa campo auto_shift_tasks (checkbox)
    project.auto_shift_tasks = request.form.get("auto_shift_tasks") == "1"

    db.session.commit()

    return redirect(url_for("project_detail", project_id=project.id))


@app.route("/projects/<int:project_id>/status/auto", methods=["POST"])
def reactivate_automatic_status(project_id: int):
    """
    Reativa o status automático de um projeto, removendo o override manual.
    """
    project = Project.query.get_or_404(project_id)
    
    # Remove override manual
    project.status_manual = False
    project.status_manual_value = None
    
    # Calcula e salva status automático
    automatic_status = calculate_automatic_status(project)
    project.status = automatic_status
    
    db.session.commit()
    
    return redirect(url_for("project_detail", project_id=project.id))


@app.route("/projects/<int:project_id>/delete", methods=["POST"])
def delete_project(project_id: int):
    """
    Exclui um projeto e todos os seus itens relacionados.
    """
    project = Project.query.get_or_404(project_id)
    db.session.delete(project)
    db.session.commit()
    return redirect(url_for("list_projects"))


@app.route("/macrostages/<int:macrostage_id>/update", methods=["POST"])
def update_macrostage(macrostage_id: int):
    """
    Atualiza o nome de uma macroetapa.
    """
    macrostage = MacroStage.query.get_or_404(macrostage_id)
    name = request.form.get("name", "").strip()

    if name:
        macrostage.name = name
        db.session.commit()

    return redirect_with_anchor("project_detail", f"macrostage-{macrostage.id}", project_id=macrostage.project.id)


@app.route("/macrostages/<int:macrostage_id>/delete", methods=["POST"])
def delete_macrostage(macrostage_id: int):
    """
    Remove uma macroetapa.
    """
    macrostage = MacroStage.query.get_or_404(macrostage_id)
    project_id = macrostage.project.id
    db.session.delete(macrostage)
    db.session.commit()

    project = Project.query.get(project_id)
    if project:
        recalculate_project(project)
        recalculate_project_status(project)
        db.session.commit()

    return redirect_with_anchor("project_detail", f"macrostage-{macrostage_id}", project_id=project_id)


@app.route("/stages/<int:stage_id>/update", methods=["POST"])
def update_stage(stage_id: int):
    """
    Atualiza as informações de uma etapa.
    """
    stage = Stage.query.get_or_404(stage_id)
    name = request.form.get("name", "").strip()
    if name:
        stage.name = name

    stage_type = (request.form.get("stage_type", "") or "não se aplica").strip()
    if stage_type not in STAGE_TYPE_CHOICES:
        stage_type = "não se aplica"

    stage.stage_type = stage_type

    if stage_type in ("robô", "sistema"):
        stage.scope = request.form.get("scope", "").strip() or None
        tools_selected = request.form.getlist("tools")
        stage.tools = ",".join(tools_selected) if tools_selected else None
        stage.other_tools = request.form.get("other_tools", "").strip() or None
    else:
        stage.scope = None
        stage.tools = None
        stage.other_tools = None

    db.session.commit()

    return redirect_with_anchor(
        "project_detail",
        f"stage-{stage.id}",
        project_id=stage.macrostage.project.id,
    )


@app.route("/stages/<int:stage_id>/delete", methods=["POST"])
def delete_stage(stage_id: int):
    """
    Remove uma etapa e suas tarefas.
    """
    stage = Stage.query.get_or_404(stage_id)
    macrostage_id = stage.macrostage.id
    project_id = stage.macrostage.project.id

    for task in list(stage.tasks):
        db.session.delete(task)

    db.session.delete(stage)
    db.session.commit()

    macrostage = MacroStage.query.get(macrostage_id)
    project = Project.query.get(project_id)

    if macrostage:
        recalculate_macrostage(macrostage)
    if project:
        recalculate_project(project)
        recalculate_project_status(project)
    db.session.commit()

    return redirect(url_for("project_detail", project_id=project_id))


@app.route("/projects/<int:project_id>/macrostages/reorder", methods=["POST"])
def reorder_macrostages(project_id: int):
    """
    Atualiza a ordem das macroetapas de um projeto específico.
    """
    project = Project.query.get_or_404(project_id)
    payload = request.get_json(silent=True) or {}
    order = payload.get("order")

    if not isinstance(order, list):
        return {"status": "error", "message": "Formato inválido"}, 400

    macrostage_map = {m.id: m for m in project.macrostages}

    position = 1
    for macrostage_id in order:
        try:
            macrostage_id = int(macrostage_id)
        except (TypeError, ValueError):
            continue

        macrostage = macrostage_map.pop(macrostage_id, None)
        if macrostage is None:
            continue

        macrostage.position = position
        position += 1

    # Aplica ordem para quaisquer macroetapas não incluídas no payload
    for macrostage in macrostage_map.values():
        macrostage.position = position
        position += 1

    db.session.commit()
    return {"status": "ok"}


@app.route("/stages/<int:macrostage_id>/reorder", methods=["POST"])
def reorder_stages(macrostage_id: int):
    """
    Atualiza a ordem das etapas dentro de uma macroetapa específica.
    """
    macrostage = MacroStage.query.get_or_404(macrostage_id)
    payload = request.get_json(silent=True) or {}
    order = payload.get("order")

    if not isinstance(order, list):
        return {"status": "error", "message": "Formato inválido"}, 400

    stage_map = {s.id: s for s in macrostage.stages}

    position = 1
    for stage_id in order:
        try:
            stage_id = int(stage_id)
        except (TypeError, ValueError):
            continue

        stage = stage_map.pop(stage_id, None)
        if stage is None:
            continue

        stage.position = position
        position += 1

    for stage in stage_map.values():
        stage.position = position
        position += 1

    db.session.commit()
    return {"status": "ok"}


@app.route("/tasks/<int:stage_id>/reorder", methods=["POST"])
def reorder_tasks(stage_id: int):
    """
    Atualiza a ordem das tarefas dentro de uma etapa específica.
    """
    stage = Stage.query.get_or_404(stage_id)
    payload = request.get_json(silent=True) or {}
    order = payload.get("order")

    if not isinstance(order, list):
        return {"status": "error", "message": "Formato inválido"}, 400

    task_map = {t.id: t for t in stage.tasks}

    position = 1
    for task_id in order:
        try:
            task_id = int(task_id)
        except (TypeError, ValueError):
            continue

        task = task_map.pop(task_id, None)
        if task is None:
            continue

        task.position = position
        position += 1

    for task in task_map.values():
        task.position = position
        position += 1

    db.session.commit()
    return {"status": "ok"}


@app.route("/macrostages/<int:macrostage_id>/tasks/reorder", methods=["POST"])
def reorder_macrostage_tasks(macrostage_id: int):
    """
    Atualiza a ordem das tarefas diretamente ligadas a uma macroetapa.
    """
    macrostage = MacroStage.query.get_or_404(macrostage_id)
    payload = request.get_json(silent=True) or {}
    order = payload.get("order")

    if not isinstance(order, list):
        return {"status": "error", "message": "Formato inválido"}, 400

    direct_tasks = {t.id: t for t in macrostage.tasks if t.stage_id is None}

    position = 1
    for task_id in order:
        try:
            task_id = int(task_id)
        except (TypeError, ValueError):
            continue

        task = direct_tasks.pop(task_id, None)
        if task is None:
            continue

        task.position = position
        position += 1

    for task in direct_tasks.values():
        task.position = position
        position += 1

    db.session.commit()
    return {"status": "ok"}


@app.route("/macrostages/<int:macrostage_id>/structure", methods=["POST"])
def set_macrostage_structure(macrostage_id: int):
    """Define se a macroetapa usará etapas ou tarefas diretas."""
    macrostage = MacroStage.query.get_or_404(macrostage_id)
    requested = (request.form.get("structure_type") or "").strip()

    anchor = f"macrostage-{macrostage.id}"

    if requested not in MACRO_STRUCTURE_CHOICES:
        return redirect_with_anchor("project_detail", anchor, project_id=macrostage.project.id)

    direct_tasks = [t for t in macrostage.tasks if t.stage_id is None]

    if requested == "stages" and direct_tasks:
        return redirect_with_anchor("project_detail", anchor, project_id=macrostage.project.id)

    if requested == "tasks" and macrostage.stages:
        return redirect_with_anchor("project_detail", anchor, project_id=macrostage.project.id)

    macrostage.structure_type = requested
    db.session.commit()
    return redirect_with_anchor("project_detail", anchor, project_id=macrostage.project.id)


@app.route("/tasks/<int:task_id>/update", methods=["POST"])
def update_task(task_id: int):
    """
    Atualiza uma tarefa, incluindo datas de início e fim.
    Se o projeto tiver auto_shift_tasks ativado e houver deslocamento,
    redireciona para tela de confirmação antes de aplicar ajustes em cadeia.
    """
    task = Task.query.get_or_404(task_id)
    stage = task.stage
    macrostage = task.macrostage
    project = macrostage.project

    name = request.form.get("name", "").strip()
    start_date_str = request.form.get("start_date", "").strip()
    end_date_str = request.form.get("end_date", "").strip()

    if name:
        task.name = name

    # Armazena valores antigos ANTES de atualizar
    old_start_date = task.start_date
    old_end_date = task.end_date

    # Conversão das strings de data para objetos date
    new_start_date = parse_date_field(start_date_str)
    new_end_date = parse_date_field(end_date_str)

    # Validação das datas antes de atualizar
    is_valid, error_message = validate_task_dates(new_start_date, new_end_date)
    if not is_valid:
        project_id = project.id
        if stage:
            anchor = f"stage-{stage.id}"
        else:
            anchor = f"macrostage-{macrostage.id}"
        url = url_for("project_detail", project_id=project_id, error=quote(error_message))
        if anchor:
            url = f"{url}#{anchor}"
        return redirect(url)

    # Atualiza os valores na tarefa (ainda não salva no banco)
    task.start_date = new_start_date
    task.end_date = new_end_date

    # Verifica se o projeto tem ajuste automático ativado
    if project.auto_shift_tasks:
        # Calcula delta de deslocamento
        delta_days, reference_start_date = calculate_task_shift_delta(
            old_start_date, old_end_date, new_start_date, new_end_date
        )
        
        # Se houver deslocamento, verifica tarefas subsequentes
        if delta_days != 0 and reference_start_date is not None:
            subsequent_tasks = find_subsequent_tasks(
                project.id, reference_start_date, task.id
            )
            
            # Se houver tarefas subsequentes, redireciona para confirmação
            if subsequent_tasks:
                # Salva a alteração da tarefa editada primeiro
                try:
                    db.session.commit()
                except ValueError as e:
                    db.session.rollback()
                    project_id = project.id
                    if stage:
                        anchor = f"stage-{stage.id}"
                    else:
                        anchor = f"macrostage-{macrostage.id}"
                    url = url_for("project_detail", project_id=project_id, error=quote(str(e)))
                    if anchor:
                        url = f"{url}#{anchor}"
                    return redirect(url)
                
                # Redireciona para tela de confirmação
                return redirect(url_for(
                    "confirm_task_shift",
                    task_id=task.id,
                    delta_days=delta_days,
                    old_start_date=reference_start_date.isoformat()
                ))

    # Se não houver ajuste em cadeia, salva normalmente e recalcula datas agregadas
    try:
        if stage:
            recalculate_all_from_stage(stage)
            project_id = stage.macrostage.project.id
            anchor = f"stage-{stage.id}"
        else:
            recalculate_all_from_macrostage(macrostage)
            project_id = macrostage.project.id
            anchor = f"macrostage-{macrostage.id}"
    except ValueError as e:
        db.session.rollback()
        project_id = macrostage.project.id
        if stage:
            anchor = f"stage-{stage.id}"
        else:
            anchor = f"macrostage-{macrostage.id}"
        url = url_for("project_detail", project_id=project_id, error=quote(str(e)))
        if anchor:
            url = f"{url}#{anchor}"
        return redirect(url)

    return redirect_with_anchor("project_detail", anchor, project_id=project_id)


@app.route("/tasks/<int:task_id>/confirm_shift", methods=["GET"])
def confirm_task_shift(task_id: int):
    """
    Exibe tela de confirmação para ajuste automático de tarefas em cadeia.
    Recebe delta_days e old_start_date via query string.
    """
    task = Task.query.get_or_404(task_id)
    project = task.macrostage.project
    
    # Valida que o projeto tem auto_shift_tasks ativado
    if not project.auto_shift_tasks:
        return redirect(url_for("project_detail", project_id=project.id))
    
    # Recebe parâmetros da query string
    delta_days_str = request.args.get("delta_days", "").strip()
    old_start_date_str = request.args.get("old_start_date", "").strip()
    
    try:
        delta_days = int(delta_days_str)
        old_start_date = datetime.strptime(old_start_date_str, "%Y-%m-%d").date() if old_start_date_str else None
    except (ValueError, TypeError):
        return redirect(url_for("project_detail", project_id=project.id))
    
    if delta_days == 0 or old_start_date is None:
        return redirect(url_for("project_detail", project_id=project.id))
    
    # Recalcula lista de tarefas subsequentes (garantindo consistência)
    subsequent_tasks = find_subsequent_tasks(project.id, old_start_date, task.id)
    
    # Prepara dados das tarefas impactadas
    affected_tasks = []
    for subsequent_task in subsequent_tasks:
        old_start = subsequent_task.start_date
        old_end = subsequent_task.end_date
        
        # Calcula novas datas
        new_start = (old_start + timedelta(days=delta_days)) if old_start else None
        new_end = (old_end + timedelta(days=delta_days)) if old_end else None
        
        # Obtém nome da macroetapa e etapa
        macrostage_name = subsequent_task.macrostage.name
        stage_name = subsequent_task.stage.name if subsequent_task.stage else None
        
        affected_tasks.append({
            'task': subsequent_task,
            'macrostage_name': macrostage_name,
            'stage_name': stage_name,
            'old_start_date': old_start,
            'old_end_date': old_end,
            'new_start_date': new_start,
            'new_end_date': new_end,
        })
    
    return render_template(
        "confirm_task_shift.html",
        task=task,
        project=project,
        delta_days=delta_days,
        old_start_date=old_start_date,
        affected_tasks=affected_tasks,
    )


@app.route("/tasks/<int:task_id>/apply_shift", methods=["POST"])
def apply_task_shift(task_id: int):
    """
    Aplica ajustes em cadeia nas tarefas subsequentes após confirmação do usuário.
    """
    task = Task.query.get_or_404(task_id)
    project = task.macrostage.project
    
    # Valida que o projeto tem auto_shift_tasks ativado
    if not project.auto_shift_tasks:
        flash("Ajuste automático não está ativado para este projeto.", "error")
        return redirect(url_for("project_detail", project_id=project.id))
    
    # Recebe parâmetros do formulário
    delta_days_str = request.form.get("delta_days", "").strip()
    old_start_date_str = request.form.get("old_start_date", "").strip()
    
    try:
        delta_days = int(delta_days_str)
        old_start_date = datetime.strptime(old_start_date_str, "%Y-%m-%d").date() if old_start_date_str else None
    except (ValueError, TypeError):
        flash("Parâmetros inválidos.", "error")
        return redirect(url_for("project_detail", project_id=project.id))
    
    if delta_days == 0 or old_start_date is None:
        flash("Nenhum ajuste a ser aplicado.", "error")
        return redirect(url_for("project_detail", project_id=project.id))
    
    # Recalcula lista de tarefas subsequentes (garantindo consistência)
    subsequent_tasks = find_subsequent_tasks(project.id, old_start_date, task.id)
    
    if not subsequent_tasks:
        flash("Nenhuma tarefa subsequente foi encontrada.", "info")
        return redirect(url_for("project_detail", project_id=project.id))
    
    # Aplica ajustes em transação
    try:
        # A tarefa editada já foi salva na rota update_task
        # Agora aplicamos ajustes nas tarefas subsequentes
        # IMPORTANTE: Atualizamos end_date primeiro sempre para evitar conflito
        # com o validador do modelo que verifica start_date <= end_date
        # Quando postergando (delta_days > 0), end_date aumenta primeiro, então start_date pode aumentar depois
        # Quando antecipando (delta_days < 0), end_date diminui primeiro, então start_date pode diminuir depois
        for subsequent_task in subsequent_tasks:
            # Calcula as novas datas primeiro
            new_start_date = None
            new_end_date = None
            
            if subsequent_task.start_date:
                new_start_date = subsequent_task.start_date + timedelta(days=delta_days)
            if subsequent_task.end_date:
                new_end_date = subsequent_task.end_date + timedelta(days=delta_days)
            
            # Valida que as novas datas são válidas (start_date <= end_date)
            if new_start_date is not None and new_end_date is not None:
                if new_start_date > new_end_date:
                    raise ValueError(
                        f"Erro ao deslocar tarefa '{subsequent_task.name}': "
                        f"a nova data de início ({new_start_date.strftime('%d/%m/%Y')}) "
                        f"seria maior que a nova data de fim ({new_end_date.strftime('%d/%m/%Y')})."
                    )
            
            # Atualiza na ordem correta dependendo se está postergando ou antecipando
            # para evitar conflito com o validador do modelo que verifica start_date <= end_date
            if delta_days > 0:
                # POSTERGANDO: atualiza end_date primeiro (aumenta)
                # Quando start_date for atualizado depois, end_date já estará maior
                if new_end_date is not None:
                    subsequent_task.end_date = new_end_date
                if new_start_date is not None:
                    subsequent_task.start_date = new_start_date
            else:
                # ANTECIPANDO: atualiza start_date primeiro (diminui)
                # Quando end_date for atualizado depois, start_date já estará menor
                if new_start_date is not None:
                    subsequent_task.start_date = new_start_date
                if new_end_date is not None:
                    subsequent_task.end_date = new_end_date
        
        # Recalcula datas agregadas para todas as etapas/macroetapas afetadas
        # Coleta todas as etapas e macroetapas únicas das tarefas afetadas
        affected_stages = set()
        affected_macrostages = set()
        
        for subsequent_task in subsequent_tasks:
            if subsequent_task.stage:
                affected_stages.add(subsequent_task.stage)
            affected_macrostages.add(subsequent_task.macrostage)
        
        # Recalcula etapas afetadas
        for stage in affected_stages:
            recalculate_stage(stage)
        
        # Recalcula macroetapas afetadas
        for macrostage in affected_macrostages:
            recalculate_macrostage(macrostage)
        
        # Recalcula projeto
        recalculate_project(project)
        recalculate_project_status(project)
        
        # Commit de todas as alterações
        db.session.commit()
        
        # Mensagem de sucesso
        if delta_days > 0:
            flash(f"Cronograma atualizado. {len(subsequent_tasks)} tarefa{'s' if len(subsequent_tasks) != 1 else ''} foram postergada{'s' if len(subsequent_tasks) != 1 else ''} em {delta_days} dia{'s' if delta_days != 1 else ''}.", "success")
        else:
            flash(f"Cronograma atualizado. {len(subsequent_tasks)} tarefa{'s' if len(subsequent_tasks) != 1 else ''} foram antecipada{'s' if len(subsequent_tasks) != 1 else ''} em {-delta_days} dia{'s' if -delta_days != 1 else ''}.", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao aplicar ajustes: {str(e)}", "error")
        return redirect(url_for("project_detail", project_id=project.id))
    
    return redirect(url_for("project_detail", project_id=project.id))


@app.route("/tasks/<int:task_id>/delete", methods=["POST"])
def delete_task(task_id: int):
    """
    Remove uma tarefa e recalcula as datas associadas.
    """
    task = Task.query.get_or_404(task_id)
    stage_id = task.stage_id
    macrostage_id = task.macrostage_id
    project_id = task.macrostage.project.id

    db.session.delete(task)
    db.session.commit()

    stage = Stage.query.get(stage_id)
    if stage:
        recalculate_all_from_stage(stage)
        anchor = f"stage-{stage.id}"
    else:
        macrostage = MacroStage.query.get(macrostage_id)
        if macrostage:
            recalculate_all_from_macrostage(macrostage)
            anchor = f"macrostage-{macrostage.id}"
        else:
            anchor = ""

    return redirect_with_anchor("project_detail", anchor, project_id=project_id)


@app.route("/tasks/<int:task_id>/weekly_updates/create", methods=["POST"])
def create_weekly_update(task_id: int):
    """
    Cria uma nova atualização semanal para a tarefa informada.
    """
    task = Task.query.get_or_404(task_id)
    content = request.form.get("content", "").strip()
    update_date_str = request.form.get("update_date", "").strip()

    if not content:
        project_id, anchor = task_parent_context(task)
        return redirect_with_anchor("project_detail", anchor, project_id=project_id)

    update = WeeklyUpdate(
        task=task,
        content=content,
        update_date=parse_date_field(update_date_str),
    )
    db.session.add(update)
    db.session.commit()

    project_id, anchor = task_parent_context(task)
    return redirect_with_anchor("project_detail", anchor, project_id=project_id)


@app.route("/weekly_updates/<int:update_id>/update", methods=["POST"])
def update_weekly_update(update_id: int):
    """
    Atualiza o conteúdo ou a data de uma atualização semanal.
    """
    weekly_update = WeeklyUpdate.query.get_or_404(update_id)

    content = request.form.get("content", "").strip()
    update_date_str = request.form.get("update_date", "").strip()

    if content:
        weekly_update.content = content
    weekly_update.update_date = parse_date_field(update_date_str)

    db.session.commit()

    project_id, anchor = task_parent_context(weekly_update.task)
    return redirect_with_anchor("project_detail", anchor, project_id=project_id)


@app.route("/weekly_updates/<int:update_id>/delete", methods=["POST"])
def delete_weekly_update(update_id: int):
    """
    Remove uma atualização semanal específica.
    """
    weekly_update = WeeklyUpdate.query.get_or_404(update_id)
    project_id, anchor = task_parent_context(weekly_update.task)

    db.session.delete(weekly_update)
    db.session.commit()

    return redirect_with_anchor("project_detail", anchor, project_id=project_id)


if __name__ == "__main__":
    # Cria as tabelas antes de iniciar o servidor (substitui o before_first_request)
    with app.app_context():
        db.create_all()

    # Configuração para funcionar tanto localmente quanto no Render
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
